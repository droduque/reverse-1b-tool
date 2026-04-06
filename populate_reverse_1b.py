"""
SVN Rock — Phase 1: Populate Reverse 1B from 1A Proforma
=========================================================
Reads a 1A proforma Excel file, copies the Reverse 1B template,
and writes the 1A data into the correct cells across 3 sheets.

Usage:
    python populate_reverse_1b.py reference/1A_Birchmount_2240.xlsx
    python populate_reverse_1b.py reference/1A_490_St_Clair.xls

The tool will prompt for municipality selection (for DC rates)
and building type (mid-rise vs high-rise).
"""

import sys
import os
import shutil
import math
import re
import json
from datetime import date
import openpyxl
from data_freshness import get_data_sources_log_block, get_alerts, get_data_sources_footer

# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "reference", "REVERSE_1B_Template.xlsx")
DC_RATES_PATH = os.path.join(os.path.dirname(__file__), "reference", "Municipalities & DC's & Prop Taxes for Proforma Testing.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")

# GFA efficiency ratio — derived from Birchmount template (128,853 / 146,346)
GFA_EFFICIENCY = 0.88

# Amenity SF per unit — derived from Birchmount (3,703 / 170 = 21.8)
AMENITY_SF_PER_UNIT = 22

# Parking SF per space — industry standard for underground
PARKING_SF_PER_SPACE = 350

# High-rise threshold — 7+ storeys per Noor (2026-03-09)
HIGH_RISE_FLOOR_THRESHOLD = 7

# Financing programs — Sheet 6 D31:D35 (permanent take-out loan parameters)
# Each program maps to the same 5 cells: LTV, DSCR, amortization, rate, premium
# Source: Joanna's proformas (2240 Birchmount Rd ProForma v3.12)
FINANCING_PROGRAMS = {
    'cmhc_mli_100': {
        # MLI Select at 100 points — max flexibility tier
        # Premium: 3.75% base + 1.25% amort surcharge (50yr) = 5.00%
        'label': 'CMHC MLI Select (100pts)',
        'max_ltv': 0.95,
        'min_dscr': 1.1,
        'amortization': 50,
        'interest_rate': 0.037,
        'cmhc_premium': 0.05,
    },
    'cmhc_mli_50': {
        # MLI Select Energy at 50 points — energy efficiency tier
        # Premium: 5.00% base + 0.75% amort surcharge (40yr), 10% energy discount = 5.175%
        'label': 'CMHC MLI Select Energy (50pts)',
        'max_ltv': 0.95,
        'min_dscr': 1.1,
        'amortization': 40,
        'interest_rate': 0.037,
        'cmhc_premium': 0.05175,
    },
    'conventional': {
        'label': 'Conventional',
        'max_ltv': 0.75,
        'min_dscr': 1.2,
        'amortization': 25,
        'interest_rate': 0.055,
        'cmhc_premium': 0,
    },
}
# Backwards compat — old JSON files may reference these keys
FINANCING_PROGRAMS['cmhc_mli_select'] = FINANCING_PROGRAMS['cmhc_mli_100']
FINANCING_PROGRAMS['cmhc_standard'] = FINANCING_PROGRAMS['cmhc_mli_100']

# Unit type grouping patterns — match labels to 3 template rows
# Order matters: checked top-to-bottom, first match wins
UNIT_GROUP_PATTERNS = [
    ("1 Bed", ["studio", "bachelor", "1 bed", "1bed", "one bed"]),
    ("2 Bed", ["2 bed", "2bed", "two bed"]),
    ("3 Bed", ["3 bed", "3bed", "three bed", "4 bed", "4bed"]),
]

# ---------------------------------------------------------------------------
# DC RATE LOOKUP — loads municipality rates from Kanen's spreadsheet
# ---------------------------------------------------------------------------

def load_dc_rates():
    """
    Parse Kanen's DC spreadsheet into a list of municipality entries.
    Each entry: {name, rates: {1bed, 2bed, 3bed}, notes}

    The spreadsheet has varying column usage per city. We consolidate
    into 3 categories to match the current 3-row template:
      - 1bed: best available from cols C (bachelor), D (1bed), E (1bed), F (1bed+den)
      - 2bed: best available from cols G (2bed), H (2bed+)
      - 3bed: best available from cols I (3bed), J (3bed+), K (4bed+)
    """
    if not os.path.exists(DC_RATES_PATH):
        print(f"Warning: DC rates file not found at {DC_RATES_PATH}")
        return []

    wb = openpyxl.load_workbook(DC_RATES_PATH, data_only=True)
    ws = wb['2026 DC\'s']

    municipalities = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=12, values_only=False):
        city_cell = row[1]  # Column B
        if not city_cell.value:
            continue

        city_name = str(city_cell.value).strip()

        # Read all rate columns — some cities only populate certain ones
        def num(cell):
            v = cell.value
            if v is None:
                return None
            if isinstance(v, str):
                # Handle text-formatted numbers like "$39,088.07"
                cleaned = v.replace('$', '').replace(',', '').strip()
                try:
                    return float(cleaned)
                except ValueError:
                    return None
            return float(v)

        col_c = num(row[2])   # Bachelor/Studio
        col_d = num(row[3])   # 1 Bed & Bachelor
        col_e = num(row[4])   # 1 Bed
        col_f = num(row[5])   # 1 Bed + Den
        col_g = num(row[6])   # 2 Bed
        col_h = num(row[7])   # 2 Bed +
        col_i = num(row[8])   # 3 Bed
        col_j = num(row[9])   # 3 Bed +
        col_k = num(row[10])  # 4 Bed +
        notes = str(row[11].value or '').strip()

        # Consolidate into 3 categories using fallback chains
        # 1-bed: prefer D (most common), fall back to E, C, F
        rate_1bed = col_d or col_e or col_c or col_f
        # 2-bed: prefer G, fall back to H
        rate_2bed = col_g or col_h
        # 3-bed: prefer J, fall back to I, then K, then 2-bed rate
        rate_3bed = col_j or col_i or col_k or rate_2bed

        # If 2-bed is missing but 3-bed exists, use 3-bed as fallback
        if rate_2bed is None and rate_3bed is not None:
            rate_2bed = rate_3bed
        # If still missing, use 1-bed as last resort
        if rate_2bed is None:
            rate_2bed = rate_1bed
        if rate_3bed is None:
            rate_3bed = rate_2bed

        if rate_1bed is None:
            continue  # Skip rows with no usable data

        municipalities.append({
            'name': city_name,
            'rates': {
                '1bed': round(rate_1bed),
                '2bed': round(rate_2bed),
                '3bed': round(rate_3bed),
            },
            # Keep the granular rates for when Noor expands to 7+ rows
            'rates_detailed': {
                'bachelor': col_c,
                '1bed_bachelor': col_d,
                '1bed': col_e,
                '1bed_den': col_f,
                '2bed': col_g,
                '2bed_plus': col_h,
                '3bed': col_i,
                '3bed_plus': col_j,
                '4bed_plus': col_k,
            },
            'notes': notes,
            'tax_rate': None,  # filled below from Prop Tax sheet
        })

    # --- Load property tax rates from "Prop Tax Rates" sheet ---
    # Match by city name to the municipalities we already loaded.
    # Use "New Multi-Residential" (NT) rate — that's what new builds pay.
    # Prefer 2026 rate (col D), fall back to 2025 (col C).
    if 'Prop Tax Rates' in wb.sheetnames:
        ws_tax = wb['Prop Tax Rates']
        # Build a lookup: normalized city name -> tax rate
        tax_lookup = {}
        for row in ws_tax.iter_rows(min_row=4, max_row=ws_tax.max_row, max_col=7, values_only=False):
            city_cell = row[1]  # Column B
            if not city_cell.value:
                continue
            city_name_tax = str(city_cell.value).strip()
            # Col D = NT 2026, Col C = NT 2025, Col F = MT 2026, Col E = MT 2025
            # Prefer New Multi-Res (NT) — what new builds pay.
            # Fall back to Multi-Res (MT) if NT not available.
            nt_2026 = row[3].value if row[3].value and isinstance(row[3].value, (int, float)) else None
            nt_2025 = row[2].value if row[2].value and isinstance(row[2].value, (int, float)) else None
            mt_2026 = row[5].value if row[5].value and isinstance(row[5].value, (int, float)) else None
            mt_2025 = row[4].value if row[4].value and isinstance(row[4].value, (int, float)) else None
            rate = nt_2026 or nt_2025 or mt_2026 or mt_2025
            if rate:
                # Normalize: "Toronto, ON" -> "TORONTO"
                key = city_name_tax.split(',')[0].strip().upper()
                tax_lookup[key] = rate

        # Attach tax rates to municipalities by matching city names
        for m in municipalities:
            muni_key = m['name'].split(',')[0].split('(')[0].strip().upper()
            if muni_key in tax_lookup:
                m['tax_rate'] = tax_lookup[muni_key]
            else:
                # Try partial matches (e.g., "Hamilton" matches "Hamilton, ON")
                for tax_key, rate in tax_lookup.items():
                    if tax_key in muni_key or muni_key in tax_key:
                        m['tax_rate'] = rate
                        break

    # Sort alphabetically by city name for cleaner dropdown display
    municipalities.sort(key=lambda m: m['name'].split(',')[0].split('(')[0].strip().upper())

    return municipalities


def select_municipality(municipalities):
    """
    Display a numbered list of municipalities and let the user pick one.
    Returns the selected municipality dict, or None if skipped.
    """
    if not municipalities:
        print("\nNo DC rates available. Using defaults.")
        return None

    print(f"\n{'='*60}")
    print("SELECT MUNICIPALITY (for development charges)")
    print(f"{'='*60}")
    for i, m in enumerate(municipalities, 1):
        notes = f"  — {m['notes']}" if m['notes'] else ""
        print(f"  {i:>2}. {m['name']}{notes}")
    print(f"  {len(municipalities)+1:>2}. Skip (use no DC rates)")

    while True:
        try:
            choice = input(f"\nEnter number (1-{len(municipalities)+1}): ").strip()
            idx = int(choice)
            if idx == len(municipalities) + 1:
                return None
            if 1 <= idx <= len(municipalities):
                selected = municipalities[idx - 1]
                print(f"\n  Selected: {selected['name']}")
                print(f"  DC rates: 1-Bed ${selected['rates']['1bed']:,} | "
                      f"2-Bed ${selected['rates']['2bed']:,} | "
                      f"3-Bed ${selected['rates']['3bed']:,}")
                return selected
        except (ValueError, IndexError):
            pass
        print(f"  Please enter a number between 1 and {len(municipalities)+1}")


def select_building_type(total_units):
    """
    Ask user for building type (mid-rise vs high-rise).
    Estimates floor count to suggest the likely answer.
    """
    est_floors = math.ceil(total_units / 12)

    print(f"\n{'='*60}")
    print("SELECT BUILDING TYPE")
    print(f"{'='*60}")
    print(f"  Estimated floors: ~{est_floors} (based on {total_units} units)")
    if est_floors >= HIGH_RISE_FLOOR_THRESHOLD:
        print(f"  Likely HIGH-RISE (7+ storeys)")
    else:
        print(f"  Likely MID-RISE (under 7 storeys)")

    print(f"\n  1. Mid-Rise (up to 6 storeys)")
    print(f"  2. High-Rise (7+ storeys)")

    while True:
        try:
            choice = input(f"\nEnter number (1-2): ").strip()
            idx = int(choice)
            if idx == 1:
                print("  Selected: Mid-Rise")
                return 'mid-rise'
            elif idx == 2:
                print("  Selected: High-Rise")
                return 'high-rise'
        except (ValueError, IndexError):
            pass
        print("  Please enter 1 or 2")


# Default operating expense values (used only if not found in 1A)
DEFAULT_RM_PER_UNIT = 1050
DEFAULT_STAFFING_PER_UNIT = 1200
DEFAULT_INSURANCE_PER_UNIT = 450
DEFAULT_MARKETING_PER_UNIT = 300
DEFAULT_GA_PER_UNIT = 250
DEFAULT_UTILITIES_PSF = 11
DEFAULT_RESERVE_PCT = 0.02


# ---------------------------------------------------------------------------
# 1A PARSER — reads any 1A proforma dynamically
# ---------------------------------------------------------------------------

def parse_1a(filepath):
    """
    Parse a 1A proforma file (.xlsx or .xls) and return a dict of all
    extracted data. Finds sections by scanning for landmark strings,
    not hardcoded row numbers.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        return _parse_1a_xls(filepath)
    else:
        return _parse_1a_xlsx(filepath)


def _parse_1a_xlsx(filepath):
    """Parse .xlsx format using openpyxl."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active  # Use the first/active sheet

    # Build a simple cell reader
    def cell(row, col_letter):
        return ws[f"{col_letter}{row}"].value

    return _parse_sheet(cell, ws.max_row)


def _parse_1a_xls(filepath):
    """Parse .xls format using xlrd, falling back to LibreOffice conversion
    if xlrd can't handle the file (some .xls files have corrupt shared string
    tables that crash xlrd but open fine in Excel/LibreOffice)."""
    import xlrd
    try:
        wb = xlrd.open_workbook(filepath)
    except Exception as e:
        # xlrd failed — try converting to .xlsx via LibreOffice
        print(f"  [xlrd] Failed to open .xls ({e}), trying LibreOffice conversion...")
        converted = _convert_xls_to_xlsx(filepath)
        if converted:
            return _parse_1a_xlsx(converted)
        raise ValueError(f"Cannot open .xls file: xlrd failed ({e}) and LibreOffice is not available")

    ws = wb.sheet_by_index(0)

    # xlrd uses 0-based indexing; our parser uses 1-based rows and letter columns
    def col_to_idx(letter):
        idx = 0
        for ch in letter:
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1  # 0-based

    def cell(row, col_letter):
        r = row - 1  # xlrd is 0-based
        c = col_to_idx(col_letter)
        if r < 0 or r >= ws.nrows or c >= ws.ncols:
            return None
        val = ws.cell_value(r, c)
        if val == '':
            return None
        return val

    return _parse_sheet(cell, ws.nrows)


def _convert_xls_to_xlsx(filepath):
    """Convert a .xls file to .xlsx using ssconvert (gnumeric) or LibreOffice.
    Returns the path to the converted file, or None if neither is available."""
    import subprocess
    import shutil
    import tempfile

    basename = os.path.splitext(os.path.basename(filepath))[0] + '.xlsx'
    outdir = tempfile.mkdtemp()
    converted = os.path.join(outdir, basename)

    # Try ssconvert first (gnumeric, ~37MB on Railway vs ~400MB for LibreOffice)
    ssconvert = shutil.which('ssconvert')
    if ssconvert:
        try:
            proc = subprocess.run(
                [ssconvert, filepath, converted],
                capture_output=True, timeout=30
            )
            if proc.returncode == 0 and os.path.exists(converted) and os.path.getsize(converted) > 0:
                print(f"  [xlrd] Converted to .xlsx via ssconvert: {converted}")
                return converted
        except Exception:
            pass

    # Fall back to LibreOffice (available locally on Mac)
    soffice = shutil.which('soffice')
    if not soffice:
        for path in ['/Applications/LibreOffice.app/Contents/MacOS/soffice',
                     '/opt/homebrew/bin/soffice', '/usr/bin/soffice',
                     '/usr/local/bin/soffice']:
            if os.path.exists(path):
                soffice = path
                break
    if not soffice:
        return None

    try:
        proc = subprocess.run(
            [soffice, '--headless', '--calc', '--convert-to', 'xlsx',
             '--outdir', outdir, filepath],
            capture_output=True, timeout=30
        )
        if proc.returncode != 0:
            return None
        if os.path.exists(converted) and os.path.getsize(converted) > 0:
            print(f"  [xlrd] Converted to .xlsx via LibreOffice: {converted}")
            return converted
        return None
    except Exception:
        return None


def _parse_sheet(cell, max_row):
    """
    Core parser logic — works with any cell accessor function.
    Scans for landmark strings to find each section dynamically.
    """
    data = {}
    parse_warnings = []  # track which sections were found vs defaulted

    # --- TITLE & ADDRESS ---
    # Title is in E2 or F2 — try both
    title = cell(2, 'E') or cell(2, 'F') or ""
    data['title'] = title

    # Extract address from title: "Estimated Stabilized Value - ADDRESS"
    address = ""
    if " - " in str(title):
        address = str(title).split(" - ", 1)[1].strip()
    data['address'] = address

    # --- Find TOTAL/AVG row to locate unit mix ---
    total_row = None
    for r in range(1, min(max_row + 1, 50)):
        val = cell(r, 'D')
        if val and "TOTAL" in str(val).upper() and "AVG" in str(val).upper():
            total_row = r
            break

    if total_row is None:
        raise ValueError("Could not find TOTAL/AVG row in 1A proforma")

    # --- UNIT MIX ---
    # Unit types are the rows between the column headers and the TOTAL row.
    # Headers are 2 rows before the first unit type (typically row 5 for headers,
    # row 6 for first unit). Scan backwards from total_row to find unit rows.
    unit_types = []
    for r in range(total_row - 1, 0, -1):
        label = cell(r, 'D')
        count = cell(r, 'F')
        # A unit row has a label in D and a numeric count in F
        if label and count and isinstance(count, (int, float)) and count > 0:
            sf = cell(r, 'E')
            rent = cell(r, 'I')
            unit_types.insert(0, {
                'label': str(label).strip(),
                'sf': float(sf) if sf else 0,
                'count': int(count),
                'rent': float(rent) if rent else 0,
            })
        else:
            # Hit a non-unit row (header or blank) — stop scanning
            break

    data['unit_types'] = unit_types
    data['total_units'] = sum(u['count'] for u in unit_types)
    data['total_rentable_sf'] = sum(u['sf'] * u['count'] for u in unit_types)

    # --- Find OPERATING REVENUES section ---
    rev_start = None
    for r in range(total_row, min(max_row + 1, 80)):
        val = cell(r, 'E')
        if val and "ESTIMATED OPERATING REVENUES" in str(val).upper():
            rev_start = r
            break

    # Scan revenue section for specific items by label
    data['parking_underground'] = {'spaces': 0, 'fee': 0}
    data['parking_visitor'] = {'spaces': 0, 'fee': 0}
    data['parking_retail'] = {'spaces': 0, 'fee': 0}
    data['storage'] = {'count': 0, 'fee': 0}
    data['submetering'] = {'count': 0, 'fee': 0}
    data['vacancy_rate'] = 0.03  # default
    data['commercial'] = {'sf': 0, 'rate': 0}
    data['commercial_vacancy'] = 0

    if rev_start:
        for r in range(rev_start + 1, rev_start + 20):
            label = str(cell(r, 'E') or '').upper().strip()
            f_val = cell(r, 'F')
            g_val = cell(r, 'G')

            # Stop if we've hit the expenses section
            if 'ESTIMATED OPERATING EXPENSES' in label:
                break

            # Skip rows that are totals, subtotals, headers, or vacancies
            # before checking parking — "Vacancies (Rent & Parking)" contains
            # "PARKING" and would falsely match the underground condition
            if 'TOTAL' in label or 'SUB-TOTAL' in label:
                continue

            if 'VACANC' in label:
                # Vacancy rows — check if residential or commercial
                if 'COMMERCIAL' in label:
                    if f_val and isinstance(f_val, (int, float)):
                        data['commercial_vacancy'] = float(f_val)
                else:
                    if f_val and isinstance(f_val, (int, float)):
                        data['vacancy_rate'] = float(f_val)
            elif 'UNDERGROUND' in label or ('PARKING' in label and 'VISITOR' not in label
                    and 'RETAIL' not in label and 'SURFACE' not in label):
                data['parking_underground'] = {
                    'spaces': int(f_val) if isinstance(f_val, (int, float)) else 0,
                    'fee': float(g_val) if isinstance(g_val, (int, float)) else 0,
                }
            elif 'VISITOR' in label:
                data['parking_visitor'] = {
                    'spaces': int(f_val) if isinstance(f_val, (int, float)) else 0,
                    'fee': float(g_val) if isinstance(g_val, (int, float)) else 0,
                }
            elif 'RETAIL' in label and 'COMMERCIAL' not in label:
                data['parking_retail'] = {
                    'spaces': int(f_val) if isinstance(f_val, (int, float)) else 0,
                    'fee': float(g_val) if isinstance(g_val, (int, float)) else 0,
                }
            elif 'STORAGE' in label or 'LOCKER' in label:
                data['storage'] = {
                    'count': int(f_val) if isinstance(f_val, (int, float)) else 0,
                    'fee': float(g_val) if isinstance(g_val, (int, float)) else 0,
                }
            elif 'SUBMETER' in label:
                data['submetering'] = {
                    'count': int(f_val) if isinstance(f_val, (int, float)) else data['total_units'],
                    'fee': float(g_val) if isinstance(g_val, (int, float)) else 20,
                }
            elif 'COMMERCIAL' in label or 'NET COMMERCIAL' in label:
                data['commercial'] = {
                    'sf': int(f_val) if isinstance(f_val, (int, float)) else 0,
                    'rate': float(g_val) if isinstance(g_val, (int, float)) else 0,
                }

    # --- Find OPERATING EXPENSES section ---
    exp_start = None
    for r in range(rev_start or total_row, min(max_row + 1, 100)):
        val = cell(r, 'E')
        if val and "ESTIMATED OPERATING EXPENSES" in str(val).upper():
            exp_start = r
            break

    data['expenses'] = {}
    data['mgmt_fee_pct'] = 0.0425  # default
    data['tax_rate'] = 0
    data['assessed_value'] = 0
    data['reserve_pct'] = DEFAULT_RESERVE_PCT

    if exp_start:
        for r in range(exp_start + 1, exp_start + 15):
            label = str(cell(r, 'E') or '').upper().strip()
            h_val = cell(r, 'H')  # annual per unit
            f_val = cell(r, 'F')
            g_val = cell(r, 'G')

            if 'UTILIT' in label:
                data['expenses']['utilities'] = float(h_val) if h_val else None
            elif 'REPAIR' in label or 'MAINTENANCE' in label:
                data['expenses']['rm'] = float(h_val) if h_val else None
            elif 'STAFF' in label:
                data['expenses']['staffing'] = float(h_val) if h_val else None
            elif 'INSURANCE' in label:
                data['expenses']['insurance'] = float(h_val) if h_val else None
            elif 'MARKETING' in label:
                data['expenses']['marketing'] = float(h_val) if h_val else None
            elif 'GENERAL' in label or 'ADMIN' in label or 'MISCELLANEOUS' in label:
                data['expenses']['ga'] = float(h_val) if h_val else None
            elif 'MANAGEMENT' in label and 'FEE' in label:
                if g_val and isinstance(g_val, (int, float)):
                    data['mgmt_fee_pct'] = float(g_val)
            elif 'PROPERTY TAX' in label or 'MUNICIPAL' in label:
                if f_val and isinstance(f_val, (int, float)):
                    data['tax_rate'] = float(f_val)
                if g_val and isinstance(g_val, (int, float)):
                    data['assessed_value'] = float(g_val)
            elif 'RESERVE' in label:
                if g_val and isinstance(g_val, (int, float)):
                    data['reserve_pct'] = float(g_val)

    # --- Find VALUATION section ---
    # "ESTIMATED VALUATION" appears in col F or G depending on the proforma version
    data['cap_rates'] = []
    for r in range(exp_start or total_row, min(max_row + 1, 120)):
        for check_col in ('F', 'G'):
            val = cell(r, check_col)
            if val and "ESTIMATED VALUATION" in str(val).upper():
                # Next 3 rows have cap rates in column H
                for cr_row in range(r + 1, r + 4):
                    cap = cell(cr_row, 'H')
                    if cap and isinstance(cap, (int, float)):
                        data['cap_rates'].append(float(cap))
                break
        if data['cap_rates']:
            break

    # --- GFA (from internal section if available) ---
    # Scan for "Building GFA" label in E column, rows 55+
    data['gfa'] = None
    for r in range(50, min(max_row + 1, 80)):
        val = cell(r, 'E')
        if val and 'BUILDING GFA' in str(val).upper():
            gfa_val = cell(r, 'F')
            if gfa_val and isinstance(gfa_val, (int, float)):
                data['gfa'] = float(gfa_val)
            break

    # --- Amenity space (from internal section if available) ---
    data['amenity_sf'] = None
    for r in range(50, min(max_row + 1, 80)):
        val = cell(r, 'E')
        if val and 'AMENITY' in str(val).upper() and 'SPACE' in str(val).upper():
            am_val = cell(r, 'F')
            if am_val and isinstance(am_val, (int, float)):
                data['amenity_sf'] = float(am_val)
            break

    # --- PARSER WARNINGS — report what was found vs defaulted ---
    sections_found = 0
    sections_total = 10

    if data.get('unit_types'):
        sections_found += 1
    else:
        parse_warnings.append('Unit mix section not found')

    if rev_start:
        sections_found += 1
    else:
        parse_warnings.append('Revenue section not found — using defaults for all revenue items')

    if exp_start:
        sections_found += 1
    else:
        parse_warnings.append('Expense section not found — using defaults for all expenses')

    if len(data.get('cap_rates', [])) >= 3:
        sections_found += 1
    else:
        parse_warnings.append(f"Found {len(data.get('cap_rates', []))} of 3 cap rates — missing values will use defaults")

    # Revenue sub-items — warn if zero (common for some projects, useful to flag)
    if data['parking_underground']['spaces'] == 0:
        sections_found += 1  # still counts as "found" — just zero
        parse_warnings.append('No underground parking found (0 spaces)')
    else:
        sections_found += 1

    if data['storage']['count'] == 0:
        parse_warnings.append('No storage lockers found (0 units)')
        sections_found += 1
    else:
        sections_found += 1

    if data['submetering']['count'] == 0 or data['submetering']['fee'] == 0:
        parse_warnings.append('No submetering revenue found')
        sections_found += 1
    else:
        sections_found += 1

    if data['commercial']['sf'] == 0:
        parse_warnings.append('No commercial space found (0 SF)')
        sections_found += 1
    else:
        sections_found += 1

    if data.get('gfa') is None:
        parse_warnings.append('GFA not found in 1A — will estimate from net rentable SF')
        sections_found += 1
    else:
        sections_found += 1

    # Tax data
    if data.get('tax_rate', 0) == 0 and data.get('assessed_value', 0) == 0:
        parse_warnings.append('No property tax data found in 1A')
        sections_found += 1
    else:
        sections_found += 1

    data['parse_warnings'] = parse_warnings
    data['sections_found'] = sections_found
    data['sections_total'] = sections_total

    return data


# ---------------------------------------------------------------------------
# IRR SOLVER — Newton-Raphson method (used by both verified and Python metrics)
# ---------------------------------------------------------------------------

def _solve_irr(cash_flows, initial_guess=0.12):
    """
    Solve for IRR using Newton-Raphson method on NPV(r) = 0.
    Returns the rate as a decimal (0.12 = 12%), or the last guess if
    convergence fails. Mirrors the JS version in presentation.html.
    """
    guess = initial_guess
    for _ in range(50):
        npv = sum(c / (1 + guess) ** y for y, c in enumerate(cash_flows))
        dnpv = sum(-y * c / (1 + guess) ** (y + 1) for y, c in enumerate(cash_flows) if y > 0)
        if abs(dnpv) < 1e-10:
            break
        next_g = guess - npv / dnpv
        if abs(next_g - guess) < 1e-8:
            guess = next_g
            break
        guess = max(-0.5, min(next_g, 2.0))
    return guess


# ---------------------------------------------------------------------------
# UNIT MIX CONSOLIDATION — groups N unit types into 3 rows
# ---------------------------------------------------------------------------

def consolidate_unit_mix(unit_types):
    """
    Groups any number of unit types into 3 categories (1-Bed, 2-Bed, 3-Bed)
    using weighted averages for SF and rent.

    Returns a list of 3 dicts: [{'label', 'sf', 'count', 'rent'}, ...]
    """
    groups = {name: [] for name, _ in UNIT_GROUP_PATTERNS}

    for unit in unit_types:
        label_lower = unit['label'].lower()
        matched = False
        for group_name, patterns in UNIT_GROUP_PATTERNS:
            if any(p in label_lower for p in patterns):
                groups[group_name].append(unit)
                matched = True
                break
        if not matched:
            # Unrecognized type — put it in 1-Bed as a catch-all
            groups["1 Bed"].append(unit)

    result = []
    for group_name, _ in UNIT_GROUP_PATTERNS:
        units_in_group = groups[group_name]
        total_count = sum(u['count'] for u in units_in_group)
        if total_count > 0:
            # Weighted average SF and rent
            weighted_sf = sum(u['sf'] * u['count'] for u in units_in_group) / total_count
            weighted_rent = sum(u['rent'] * u['count'] for u in units_in_group) / total_count
            result.append({
                'label': group_name,
                'sf': round(weighted_sf, 2),
                'count': total_count,
                'rent': round(weighted_rent, 2),
            })
        else:
            result.append({'label': group_name, 'sf': 0, 'count': 0, 'rent': 0})

    return result


# ---------------------------------------------------------------------------
# TEMPLATE WRITER — copies template and writes data to 3 sheets
# ---------------------------------------------------------------------------

def populate_template(data, output_path, municipality=None, building_type='high-rise',
                      financing_program=None, construction_months=None,
                      gfa_override=None, parking_sf_override=None,
                      construction_financing=None):
    """
    Copy the Reverse 1B template and write parsed 1A data into it.
    Uses the ZIP/XML writer to preserve all drawings, images, and formatting.
    Returns a list of log entries documenting every change.

    municipality: dict from load_dc_rates() with 'name' and 'rates' keys, or None
    building_type: 'mid-rise' or 'high-rise' — affects Altus cost guide row reference
    construction_months: override for construction duration (default: auto from unit count)
    gfa_override: user-provided GFA in SF (overrides 1A/estimated value)
    parking_sf_override: user-provided total parking SF
    construction_financing: dict with mezz/bank debt %, prime rates, margins, fees
    """
    from xml_writer import write_cell, save_workbook

    log = []

    # Consolidate unit mix if needed
    if len(data['unit_types']) <= 3:
        consolidated = data['unit_types']
        # Pad to 3 rows if fewer than 3 types
        while len(consolidated) < 3:
            consolidated.append({'label': f"{len(consolidated)+1} Bed", 'sf': 0, 'count': 0, 'rent': 0})
    else:
        consolidated = consolidate_unit_mix(data['unit_types'])
        log.append(f"CONSOLIDATION: {len(data['unit_types'])} unit types consolidated into 3 groups:")
        for orig in data['unit_types']:
            log.append(f"  {orig['label']}: {orig['count']} units, {orig['sf']} SF, ${orig['rent']}/mo")
        log.append("  Grouped as:")
        for grp in consolidated:
            log.append(f"  → {grp['label']}: {grp['count']} units, {grp['sf']} SF, ${grp['rent']}/mo")
        log.append("")

    # Collect all cell writes per sheet: {cell_ref: (value, description)}
    sheet1_writes = {}
    sheet4_writes = {}
    sheet5_writes = {}
    sheet6_writes = {}

    # Resolve financing program — default to CMHC MLI Select 100pts
    if financing_program is None:
        financing_program = FINANCING_PROGRAMS['cmhc_mli_100']

    def queue_write(target, cell_ref, value, description="", force=False):
        """Queue a cell write for later application via XML writer."""
        target[cell_ref] = (value, description, force)

    # ===================================================================
    # SHEET 1: 1A Proforma
    # ===================================================================
    log.append("=" * 60)
    log.append("SHEET 1: 1. 1A Proforma")
    log.append("=" * 60)

    # Title and address (Fran V2 template: columns shifted left by 1)
    queue_write(sheet1_writes, 'E2', "Estimated Stabilized Value - Today", "Title")
    queue_write(sheet1_writes, 'E3', data['address'], "Address from 1A title")

    # Unit mix — rows 7, 8, 9
    # If a group has 0 units (e.g., Bayview has no 1-beds), clear old template
    # values by writing empty strings instead of zeros
    for i, unit in enumerate(consolidated):
        row = 7 + i
        if unit['count'] == 0:
            # Write zeros for numeric cells so formulas (H7=E7*F7 etc.) don't #VALUE!
            queue_write(sheet1_writes, f'C{row}', '', f"Unit type {i+1} label (empty — no units)")
            queue_write(sheet1_writes, f'D{row}', 0, f"Unit type {i+1} avg SF (zero — no units)")
            queue_write(sheet1_writes, f'E{row}', 0, f"Unit type {i+1} count (zero)")
            queue_write(sheet1_writes, f'H{row}', 0, f"Unit type {i+1} monthly rent (zero — no units)")
        else:
            queue_write(sheet1_writes, f'C{row}', unit['label'], f"Unit type {i+1} label")
            queue_write(sheet1_writes, f'D{row}', unit['sf'], f"Unit type {i+1} avg SF")
            queue_write(sheet1_writes, f'E{row}', unit['count'], f"Unit type {i+1} count")
            queue_write(sheet1_writes, f'H{row}', unit['rent'], f"Unit type {i+1} monthly rent")

    # Operating revenues
    queue_write(sheet1_writes, 'E18', data['parking_underground']['spaces'], "Underground parking spaces")
    queue_write(sheet1_writes, 'F18', data['parking_underground']['fee'], "Underground parking monthly fee")
    queue_write(sheet1_writes, 'E19', data['parking_visitor']['spaces'], "Visitor parking spaces")
    queue_write(sheet1_writes, 'F19', data['parking_visitor']['fee'], "Visitor parking monthly fee")
    queue_write(sheet1_writes, 'E20', data['parking_retail']['spaces'], "Retail parking spaces")
    queue_write(sheet1_writes, 'F20', data['parking_retail']['fee'], "Retail parking monthly fee")
    queue_write(sheet1_writes, 'E21', data['storage']['count'], "Storage locker count")
    queue_write(sheet1_writes, 'F21', data['storage']['fee'], "Storage locker monthly fee")

    # Submetering — F22 has an external workbook ref formula, force overwrite
    queue_write(sheet1_writes, 'F22', data['submetering']['fee'], "Submetering monthly fee — replaced external ref", force=True)

    queue_write(sheet1_writes, 'E24', data['vacancy_rate'], "Residential vacancy rate")
    queue_write(sheet1_writes, 'E26', data['commercial']['sf'], "Commercial retail SF")
    queue_write(sheet1_writes, 'F26', data['commercial']['rate'], "Commercial retail $/SF rate")
    queue_write(sheet1_writes, 'E27', data['commercial_vacancy'], "Commercial vacancy rate")

    # Operating expenses
    queue_write(sheet1_writes, 'F37', data['mgmt_fee_pct'], "Management fee %")
    queue_write(sheet1_writes, 'E38', data['tax_rate'], "Property tax rate")
    queue_write(sheet1_writes, 'F38', data['assessed_value'], "Assessed value per unit")

    # Cap rates
    if len(data['cap_rates']) >= 3:
        queue_write(sheet1_writes, 'G46', data['cap_rates'][0], "Best case cap rate")
        queue_write(sheet1_writes, 'G47', data['cap_rates'][1], "Base case cap rate")
        queue_write(sheet1_writes, 'G48', data['cap_rates'][2], "Worst case cap rate")

    # --- Internal operating assumptions (rows 58+) ---
    log.append("")
    log.append("--- Sheet 1 Internal Section (Operating Assumptions) ---")

    # GFA — user override > 1A value > estimate
    if gfa_override:
        gfa = gfa_override
        queue_write(sheet1_writes, 'E62', gfa, f"Building GFA (user override: {gfa:,.0f} SF)")
    elif data.get('gfa'):
        gfa = data['gfa']
        queue_write(sheet1_writes, 'E62', gfa, "Building GFA (from 1A internal section)")
    else:
        gfa = round(data['total_rentable_sf'] / GFA_EFFICIENCY)
        queue_write(sheet1_writes, 'E62', gfa,
                    f"ESTIMATED: net rentable {data['total_rentable_sf']:.0f} / {GFA_EFFICIENCY} efficiency")

    # Amenity space — from 1A if available, else estimate
    amenity_sf = data.get('amenity_sf')
    if amenity_sf:
        queue_write(sheet1_writes, 'E64', amenity_sf, "Amenity space (from 1A)")
    else:
        amenity_sf = round(data['total_units'] * AMENITY_SF_PER_UNIT, -2)  # round to 100
        queue_write(sheet1_writes, 'E64', amenity_sf,
                    f"ESTIMATED: {data['total_units']} units x {AMENITY_SF_PER_UNIT} SF/unit, rounded to 100")

    # Per-unit operating costs — use 1A values if available, else defaults
    rm = data['expenses'].get('rm') or DEFAULT_RM_PER_UNIT
    queue_write(sheet1_writes, 'H93', rm,
                f"R&M per unit ({'from 1A' if data['expenses'].get('rm') else 'DEFAULT'})")

    staffing = data['expenses'].get('staffing') or DEFAULT_STAFFING_PER_UNIT
    queue_write(sheet1_writes, 'H109', staffing,
                f"Staffing per unit ({'from 1A' if data['expenses'].get('staffing') else 'DEFAULT'})")

    insurance = data['expenses'].get('insurance') or DEFAULT_INSURANCE_PER_UNIT
    queue_write(sheet1_writes, 'E117', insurance,
                f"Insurance per unit ({'from 1A' if data['expenses'].get('insurance') else 'DEFAULT'})")

    marketing = data['expenses'].get('marketing') or DEFAULT_MARKETING_PER_UNIT
    queue_write(sheet1_writes, 'E122', marketing,
                f"Marketing per unit ({'from 1A' if data['expenses'].get('marketing') else 'DEFAULT'})")

    ga = data['expenses'].get('ga') or DEFAULT_GA_PER_UNIT
    queue_write(sheet1_writes, 'E127', ga,
                f"G&A per unit ({'from 1A' if data['expenses'].get('ga') else 'DEFAULT'})")

    queue_write(sheet1_writes, 'H137', data['reserve_pct'], "Reserve for replacement %")

    # Utilities — back-calculate PSF from per-unit value
    # Formula chain: F80 ($/PSF) -> F81 (=F80*common_area) -> H80 (=ROUND(F81/units,-1))
    # Common area = GFA - net_rentable (approx)
    utilities_per_unit = data['expenses'].get('utilities')
    if utilities_per_unit:
        common_area = gfa - data['total_rentable_sf']
        if common_area > 0:
            utilities_psf = round(utilities_per_unit * data['total_units'] / common_area, 2)
        else:
            utilities_psf = DEFAULT_UTILITIES_PSF
        queue_write(sheet1_writes, 'E80', utilities_psf,
                    f"BACK-CALCULATED: ${utilities_per_unit}/unit x {data['total_units']} units "
                    f"/ {common_area:.0f} SF common area = ${utilities_psf}/PSF")
    else:
        utilities_psf = DEFAULT_UTILITIES_PSF
        queue_write(sheet1_writes, 'E80', utilities_psf, "Utilities $/PSF (DEFAULT)")
    # Store all computed values back to data dict so export_project_json()
    # uses the exact same inputs written to Excel — not stale defaults.
    data['utilities_psf'] = utilities_psf
    data['gfa'] = gfa
    data['amenity_sf'] = amenity_sf
    data['expenses']['rm'] = rm
    data['expenses']['staffing'] = staffing
    data['expenses']['insurance'] = insurance
    data['expenses']['marketing'] = marketing
    data['expenses']['ga'] = ga

    # ===================================================================
    # SHEET 4: Area Schedule
    # ===================================================================
    log.append("")
    log.append("=" * 60)
    log.append("SHEET 4: 4. Area Schedule")
    log.append("=" * 60)

    # Address
    queue_write(sheet4_writes, 'A2', data['address'], "Project address")

    # --- Section 1: Residential Units ---
    log.append("")
    log.append("--- Residential Units (from 1A) ---")
    for i, unit in enumerate(consolidated):
        row = 7 + i
        if unit['count'] == 0:
            # Clear old template data for empty unit groups
            queue_write(sheet4_writes, f'C{row}', 0, f"{unit['label']} count (empty — no units)")
            queue_write(sheet4_writes, f'D{row}', 0, f"{unit['label']} avg SF (empty — no units)")
            queue_write(sheet4_writes, f'F{row}', '', f"{unit['label']} note (empty — no units)")
        else:
            queue_write(sheet4_writes, f'C{row}', unit['count'], f"{unit['label']} count")
            # Round SF to integer for the area schedule (template uses integers)
            queue_write(sheet4_writes, f'D{row}', round(unit['sf']), f"{unit['label']} avg SF")
            pct = round(unit['count'] / data['total_units'] * 100) if data['total_units'] > 0 else 0
            queue_write(sheet4_writes, f'F{row}', f"{pct}% of total units", f"{unit['label']} note")

    # --- Section 2.1: Amenity Spaces ---
    log.append("")
    log.append("--- Amenity Spaces (ESTIMATED) ---")
    # Total amenity budget then distribute across 5 rooms
    total_amenity = amenity_sf
    # Keep the same proportions as Birchmount: 32%, 27%, 16%, 14%, 11%
    amenity_rooms = [
        ('Fitness Centre', 0.32),
        ('Multi-Purpose/Party Room', 0.27),
        ('Co-Working Space', 0.16),
        ('Games/Lounge Area', 0.14),
        ('Outdoor Terrace/BBQ Area', 0.11),
    ]
    for idx, (name, pct) in enumerate(amenity_rooms):
        row = 14 + idx
        room_sf = round(total_amenity * pct, -1)  # round to 10
        queue_write(sheet4_writes, f'C{row}', 1, f"{name} quantity")
        queue_write(sheet4_writes, f'D{row}', int(room_sf),
                    f"ESTIMATED: {pct:.0%} of {total_amenity} total amenity SF")
        queue_write(sheet4_writes, f'E{row}', int(room_sf), f"{name} total SF")

    # --- Section 2.2: Common Areas ---
    log.append("")
    log.append("--- Common Areas (ESTIMATED) ---")
    est_floors = math.ceil(data['total_units'] / 12)
    elevator_count = 2 if data['total_units'] < 100 else (3 if data['total_units'] <= 250 else 4)

    common_items = [
        # (row, label, qty, sf_each, total_override, note_for_log)
        (22, 'Main Lobby', 1, 800, 800,
         "DEFAULT: 800 SF"),
        (23, 'Corridors & Hallways', data['total_units'], 25, None,
         f"ESTIMATED: {data['total_units']} units x 25 SF/unit"),
        (24, 'Elevator Lobbies', est_floors, 60, None,
         f"ESTIMATED: {est_floors} floors (ceil({data['total_units']}/12)) x 60 SF"),
        (25, 'Stairwells (2)', 2, 600, None,
         "DEFAULT: 2 stairwells x 600 SF"),
        (26, f'Elevators ({elevator_count})', elevator_count, 60, None,
         f"ESTIMATED: {elevator_count} elevators x 60 SF"),
        (27, 'Mail/Parcel Room', 1, 300, 300,
         "DEFAULT: 300 SF"),
        (28, 'Garbage/Recycling Rooms', max(1, est_floors // 4), 200, None,
         f"ESTIMATED: {max(1, est_floors // 4)} rooms (floors/4) x 200 SF"),
        (29, 'Electrical/Mechanical Rooms', 2, 200, None,
         "DEFAULT: 2 rooms x 200 SF"),
        (30, 'Storage Lockers', data['storage']['count'] or round(data['total_units'] * 0.5), 25, None,
         f"{'From 1A' if data['storage']['count'] else 'ESTIMATED: 50% of units'}: "
         f"{data['storage']['count'] or round(data['total_units'] * 0.5)} lockers x 25 SF"),
        (31, 'Janitor/Housekeeping Closets', max(1, est_floors // 5), 40, None,
         f"ESTIMATED: {max(1, est_floors // 5)} closets (floors/5) x 40 SF"),
    ]

    for row, label, qty, sf_each, total_override, note in common_items:
        queue_write(sheet4_writes, f'C{row}', qty, note)
        queue_write(sheet4_writes, f'D{row}', sf_each, f"{label} SF each")
        # E column: write total SF — xml_writer's write_cell will skip if it's a formula
        total_sf = total_override if total_override else qty * sf_each
        queue_write(sheet4_writes, f'E{row}', total_sf, f"{label} total SF")

    # --- Section 3: Commercial ---
    log.append("")
    log.append("--- Commercial (from 1A) ---")
    commercial_sf = data['commercial']['sf']
    if commercial_sf > 0:
        queue_write(sheet4_writes, 'C35', 1, "Commercial unit count")
        queue_write(sheet4_writes, 'D35', commercial_sf, "Commercial SF from 1A")
        queue_write(sheet4_writes, 'E35', commercial_sf, "Commercial total SF")
    else:
        queue_write(sheet4_writes, 'C35', 0, "No commercial in this project")
        queue_write(sheet4_writes, 'D35', 0, "No commercial SF")
        queue_write(sheet4_writes, 'E35', 0, "No commercial total SF")

    # --- Section 4.1: Parking ---
    log.append("")
    log.append("--- Parking (counts from 1A, SF estimated or overridden) ---")
    parking_items = [
        (39, 'Underground Parking', data['parking_underground']['spaces']),
        (40, 'Visitor Parking', data['parking_visitor']['spaces']),
        (41, 'Retail Parking', data['parking_retail']['spaces']),
    ]
    total_spaces = sum(s for _, _, s in parking_items)
    # If user provided total parking SF, distribute proportionally across types
    if parking_sf_override and total_spaces > 0:
        sf_per_space = round(parking_sf_override / total_spaces)
        sf_note = f"USER OVERRIDE: {parking_sf_override:,.0f} total SF / {total_spaces} spaces = {sf_per_space} SF/space"
    else:
        sf_per_space = PARKING_SF_PER_SPACE
        sf_note = f"ESTIMATED: {PARKING_SF_PER_SPACE} SF/space (industry standard)"
    for row, label, spaces in parking_items:
        queue_write(sheet4_writes, f'C{row}', spaces, f"{label} spaces (from 1A)")
        queue_write(sheet4_writes, f'D{row}', sf_per_space, sf_note)

    # --- Section 4.2: Back of House ---
    log.append("")
    log.append("--- Back of House (ESTIMATED) ---")
    mech_sf = round(data['total_units'] * 12, -2)  # ~12 SF/unit, rounded to 100
    boh_items = [
        (45, 'Loading Dock', 500, "DEFAULT: 500 SF"),
        (46, 'Building Management Office', 200, "DEFAULT: 200 SF"),
        (47, 'Security Office', 150, "DEFAULT: 150 SF"),
        (48, 'Maintenance Workshop', 300, "DEFAULT: 300 SF"),
        (49, 'Mechanical Penthouse', mech_sf,
         f"ESTIMATED: {data['total_units']} units x 12 SF/unit = {mech_sf} SF"),
    ]
    for row, label, total_sf, note in boh_items:
        queue_write(sheet4_writes, f'C{row}', 1, f"{label} quantity")
        queue_write(sheet4_writes, f'D{row}', total_sf, note)
        queue_write(sheet4_writes, f'E{row}', total_sf, f"{label} total SF")

    # --- Verification: Target GFA ---
    log.append("")
    log.append("--- Verification ---")
    queue_write(sheet4_writes, 'E64', round(gfa),
                f"Target GFA ({'from 1A' if data.get('gfa') else 'ESTIMATED from net rentable / 0.88'})")

    # ===================================================================
    # SHEET 5: Key Assumptions (true inputs only)
    # ===================================================================
    log.append("")
    log.append("=" * 60)
    log.append("SHEET 5: 5. Key Assumptions (true inputs only)")
    log.append("=" * 60)

    queue_write(sheet5_writes, 'E12', 0, "Land purchase duration (months)")
    queue_write(sheet5_writes, 'E13', 12, "Pre-development duration (months)")

    # Construction duration: use override if provided, else auto-suggest from unit count
    # Rule of thumb: <200 units = 18mo, 200-400 = 24mo, 400+ = 30mo
    auto_construction = construction_months is None
    if auto_construction:
        total = data['total_units']
        if total <= 200:
            construction_months = 18
        elif total <= 400:
            construction_months = 24
        else:
            construction_months = 30
    queue_write(sheet5_writes, 'E14', construction_months,
                f"Construction duration ({construction_months}mo — {'auto from ' + str(data['total_units']) + ' units' if auto_construction else 'user override'})")

    queue_write(sheet5_writes, 'E16', 0, "Stabilized duration (months)")
    queue_write(sheet5_writes, 'F15', -3, "Lease-up offset (months)")
    queue_write(sheet5_writes, 'E37', 0.08, "Profit percentage (8%)")

    # Construction financing — debt stack, rates, fees
    cf = construction_financing or {}
    mezz_pct = cf.get('mezz_debt_pct', 0.15)
    mezz_prime = cf.get('mezz_prime_rate', 0.0445)
    mezz_margin = cf.get('mezz_margin', 0.035)
    bank_pct = cf.get('bank_debt_pct', 0.75)
    bank_prime = cf.get('bank_prime_rate', 0.0445)
    bank_margin = cf.get('bank_margin', 0.01)
    fin_fees = cf.get('financing_fees_pct', 0.01)
    fin_contingency = cf.get('financing_contingency_pct', 0.005)

    log.append("")
    log.append("--- Construction Financing (Sheet 5) ---")
    queue_write(sheet5_writes, 'E69', mezz_pct, f"Mezzanine debt ({mezz_pct:.0%})")
    queue_write(sheet5_writes, 'I69', mezz_prime, f"Mezzanine prime rate ({mezz_prime:.2%})")
    queue_write(sheet5_writes, 'J69', mezz_margin, f"Mezzanine margin ({mezz_margin:.2%})")
    queue_write(sheet5_writes, 'E70', bank_pct, f"Bank debt ({bank_pct:.0%})")
    queue_write(sheet5_writes, 'I70', bank_prime, f"Bank prime rate ({bank_prime:.2%})")
    queue_write(sheet5_writes, 'J70', bank_margin, f"Bank margin ({bank_margin:.2%})")
    queue_write(sheet5_writes, 'D74', fin_fees, f"Financing fees ({fin_fees:.1%})")
    queue_write(sheet5_writes, 'D75', fin_contingency, f"Financing contingency ({fin_contingency:.1%})")

    # Project start date — replace =TODAY() with 1st of current month.
    # EDATE from month-end dates (29/30/31) gets clamped by short months
    # (e.g., Feb 28), permanently breaking the date chain in Sheet 10.
    # Using the 1st avoids this because every month has a 1st.
    today = date.today()
    start_date = date(today.year, today.month, 1)
    # Excel date serial: days since 1899-12-30
    excel_serial = (start_date - date(1899, 12, 30)).days
    queue_write(sheet5_writes, 'G10', excel_serial, f"Project start date ({start_date.isoformat()}) — forces 1st of month to avoid EDATE clamping", force=True)

    # Development charges — from selected municipality or skip
    if municipality:
        dc = municipality['rates']
        dc_label = municipality['name']
        queue_write(sheet5_writes, 'R57', dc['1bed'], f"{dc_label} DC: 1-Bed — ${dc['1bed']:,}")
        queue_write(sheet5_writes, 'R58', dc['2bed'], f"{dc_label} DC: 2-Bed — ${dc['2bed']:,}")
        queue_write(sheet5_writes, 'R59', dc['3bed'], f"{dc_label} DC: 3-Bed — ${dc['3bed']:,}")
        if municipality['notes']:
            log.append(f"  DC NOTES: {municipality['notes']}")
    else:
        log.append("  DC RATES: No municipality selected — cells R57:R59 left unchanged (template defaults)")

    # Building type — affects which Altus Cost Guide row Sheet 5 should reference
    # Sheet 5 F29 currently points to "13-39 Storeys" (Sheet 13 row 7)
    # Mid-rise would be "Up to 12 Storeys" (Sheet 13 row 6)
    log.append(f"  BUILDING TYPE: {building_type}")
    if building_type == 'mid-rise':
        log.append("  *** FLAG: Building is mid-rise but Sheet 5 F29 references '13-39 Storeys'.")
        log.append("      Noor should verify and update the Altus height category if needed.")

    # ===================================================================
    # SHEET 6: Permanent Financing Parameters (C31:C35, Fran V2 shifted left)
    # ===================================================================
    log.append("")
    log.append("=" * 60)
    log.append("SHEET 6: 6. Permanent Financing")
    log.append("=" * 60)
    fp_label = financing_program.get('label', 'CMHC MLI Select')
    log.append(f"  Financing program: {fp_label}")

    queue_write(sheet6_writes, 'C31', financing_program['max_ltv'], f"Max LTV ({fp_label})")
    queue_write(sheet6_writes, 'C32', financing_program['min_dscr'], f"Min DSCR ({fp_label})")
    queue_write(sheet6_writes, 'C33', financing_program['amortization'], f"Amortization years ({fp_label})")
    queue_write(sheet6_writes, 'C34', financing_program['interest_rate'], f"Interest rate ({fp_label})")
    queue_write(sheet6_writes, 'C35', financing_program['cmhc_premium'], f"CMHC premium ({fp_label})")

    # ===================================================================
    # FLAGS FOR NOOR'S REVIEW
    # ===================================================================
    log.append("")
    log.append("=" * 60)
    log.append("FLAGS FOR NOOR'S REVIEW")
    log.append("=" * 60)
    log.append("1. ALTUS COST GUIDE COLUMNS: Sheet 5 O48/P48 reference Ottawa columns (H/I)")
    log.append("   instead of GTA columns (F/G) on the Altus Cost Guide sheet.")
    log.append("   Parking refs (O49/P49) correctly use GTA. This is a pre-existing")
    log.append("   template issue — not changed by this automation.")
    log.append("")
    log.append(f"2. ALTUS HEIGHT CATEGORY: Building type set to '{building_type}'.")
    log.append(f"   Sheet 5 F29 currently points to '13-39 Storeys' (Sheet 13 row 7).")
    log.append(f"   Estimated floor count: {est_floors}.")
    if building_type == 'mid-rise':
        log.append("   *** BUILDING IS MID-RISE — Noor must update F29 to 'Up to 12 Storeys'.")
    log.append("")
    log.append("3. EXTERNAL WORKBOOK REFERENCES: Sheet 1 rows 70-77, 92-97, 107-113,")
    log.append("   115-116, 120, 125-126, 132 reference '[2]Typical Conv.Operating Expenses'")
    log.append("   which is not included. These cells show stale Birchmount values.")
    log.append("   G22 (submetering fee) was overwritten with the 1A value.")

    # ===================================================================
    # DATA SOURCES & FRESHNESS
    # ===================================================================
    log.append("")
    log.extend(get_data_sources_log_block())

    # ===================================================================
    # SAVE — apply all queued writes via ZIP/XML writer
    # ===================================================================
    def _make_modifier(writes_dict):
        """Create a modifier function for save_workbook from a writes dict."""
        def modifier(sheet_root, shared_strings, log_entries):
            for cell_ref, entry in writes_dict.items():
                value, description = entry[0], entry[1]
                force = entry[2] if len(entry) > 2 else False
                write_cell(sheet_root, cell_ref, value, shared_strings, log_entries, description, force=force)
        return modifier

    sheet_modifications = {}
    if sheet1_writes:
        sheet_modifications['xl/worksheets/sheet1.xml'] = _make_modifier(sheet1_writes)
    if sheet4_writes:
        sheet_modifications['xl/worksheets/sheet4.xml'] = _make_modifier(sheet4_writes)
    if sheet5_writes:
        sheet_modifications['xl/worksheets/sheet5.xml'] = _make_modifier(sheet5_writes)
    if sheet6_writes:
        sheet_modifications['xl/worksheets/sheet6.xml'] = _make_modifier(sheet6_writes)

    save_workbook(TEMPLATE_PATH, output_path, sheet_modifications, log)

    return log


# ---------------------------------------------------------------------------
# REVERSE 1B RE-IMPORT — read a Noor-reviewed Excel back into JSON
# ---------------------------------------------------------------------------

def import_reverse_1b(xlsx_path):
    """
    Read a reviewed/modified Reverse 1B Excel and extract all data
    needed for the presentation tool JSON. This lets Noor review and
    adjust the auto-generated file, then re-import his verified numbers
    into the presentation tool — eliminating all screening estimates.

    Uses data_only=True so formula cells return their cached (last-saved)
    values. The file MUST have been opened and saved in Excel first.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    s1 = wb['1. 1A Proforma']
    s4 = wb['4. Area Schedule']
    s5 = wb['5. Key Assumptions']

    # --- Helper to safely read a cell value ---
    def val(sheet, cell, default=0):
        v = sheet[cell].value
        if v is None:
            return default
        return v

    def num(sheet, cell, default=0):
        v = sheet[cell].value
        if v is None:
            return default
        try:
            return float(v)
        except (ValueError, TypeError):
            return default

    # --- Project info (Sheet 1, Fran V2: columns shifted left by 1) ---
    title = str(val(s1, 'E2', ''))
    address = str(val(s1, 'E3', ''))
    # If E3 is empty, try to extract from title
    if not address.strip() and title:
        address = title.replace('Estimated Stabilized Value -', '').replace('Estimated Stabilized Value–', '').strip()

    # --- Unit mix (Sheet 1, rows 7-9) ---
    unit_types = []
    for row in range(7, 10):
        label = val(s1, f'C{row}', '')
        count = num(s1, f'E{row}')
        sf = num(s1, f'D{row}')
        rent = num(s1, f'H{row}')
        if label and count > 0:
            unit_types.append({
                'label': str(label).strip(),
                'count': int(count),
                'sf': round(sf, 1),
                'rent': round(rent, 2),
            })

    total_units = sum(u['count'] for u in unit_types)
    total_rentable_sf = sum(u['count'] * u['sf'] for u in unit_types)
    est_floors = math.ceil(total_units / 12) if total_units > 0 else 1

    # --- Areas (Sheet 1 + Sheet 4) ---
    gfa = num(s1, 'E62')
    amenity_sf = num(s1, 'E64')
    common_area_sf = max(0, gfa - total_rentable_sf) if gfa > 0 else 0

    # Parking SF
    pkg_underground = int(num(s1, 'E18'))
    pkg_visitor = int(num(s1, 'E19'))
    pkg_retail = int(num(s1, 'E20'))
    parking_sf = (pkg_underground + pkg_visitor + pkg_retail) * PARKING_SF_PER_SPACE

    # --- Parking fees ---
    pkg_underground_fee = num(s1, 'F18')
    pkg_visitor_fee = num(s1, 'F19')
    pkg_retail_fee = num(s1, 'F20')

    # --- Storage ---
    storage_count = int(num(s1, 'E21'))
    storage_fee = num(s1, 'F21')

    # --- Submetering ---
    submetering_fee = num(s1, 'F22', 20)

    # --- Commercial ---
    commercial_sf = num(s1, 'E26')
    commercial_rate = num(s1, 'F26')
    commercial_vacancy = num(s1, 'E27')

    # --- Vacancy ---
    vacancy_rate = num(s1, 'E24')

    # --- Cap rates (Sheet 1, G46:G48, Fran V2 shifted) ---
    cap_best = num(s1, 'G46', 0.0425)
    cap_base = num(s1, 'G47', 0.045)
    cap_worst = num(s1, 'G48', 0.0475)

    # --- OpEx (Sheet 1, Fran V2 shifted) ---
    mgmt_fee = num(s1, 'F37', 0.0425)
    tax_rate = num(s1, 'E38')
    assessed_value = num(s1, 'F38')
    utilities_psf = num(s1, 'E80', DEFAULT_UTILITIES_PSF)
    rm_per_unit = num(s1, 'H93', DEFAULT_RM_PER_UNIT)
    staffing_per_unit = num(s1, 'H109', DEFAULT_STAFFING_PER_UNIT)
    insurance_per_unit = num(s1, 'E117', DEFAULT_INSURANCE_PER_UNIT)
    marketing_per_unit = num(s1, 'E122', DEFAULT_MARKETING_PER_UNIT)
    ga_per_unit = num(s1, 'E127', DEFAULT_GA_PER_UNIT)
    reserve_pct = num(s1, 'H137', DEFAULT_RESERVE_PCT)

    # --- Schedule (Sheet 5) ---
    land_months = int(num(s5, 'E12'))
    predev_months = int(num(s5, 'E13', 12))
    construction_months = int(num(s5, 'E14', 18))
    stabilized_months = int(num(s5, 'E16'))
    leaseup_offset = int(num(s5, 'F15', -3))

    # --- DC rates (Sheet 5, R57:R59) ---
    dc_1bed = num(s5, 'R57')
    dc_2bed = num(s5, 'R58')
    dc_3bed = num(s5, 'R59')
    dc_total = 0
    for u in unit_types:
        label_lower = u['label'].lower()
        if '1' in label_lower:
            dc_total += u['count'] * dc_1bed
        elif '2' in label_lower:
            dc_total += u['count'] * dc_2bed
        else:
            dc_total += u['count'] * dc_3bed

    # --- Profit pct (Sheet 5 E37) ---
    profit_pct = num(s5, 'E37', 0.08)

    # --- Construction cost PSF ---
    # Sheet 5 G48 has hard cost per rentable SF (formula cell — cached value)
    # Sheet 5 F48 has total hard construction cost
    construction_cost_psf = num(s5, 'G48')
    if construction_cost_psf <= 0 or construction_cost_psf > 2000:
        # Fallback: derive from total / GFA, or use baseline
        total_hard = num(s5, 'F48')
        if total_hard > 0 and gfa > 0:
            construction_cost_psf = total_hard / gfa
        else:
            construction_cost_psf = 453  # Birchmount baseline

    # --- Financing (Sheet 6 if it exists, otherwise defaults) ---
    # Permanent loan parameters in column C, rows 31-35 (Fran V2 shifted)
    # Defaults match CMHC MLI Select 100pts (Joanna's standard proforma)
    try:
        s6 = wb['6. Debt Stack & Financing']
        perm_ltv = num(s6, 'C31', 0.95)
        perm_dscr = num(s6, 'C32', 1.1)
        perm_rate = num(s6, 'C34', 0.037)
        perm_term = int(num(s6, 'C33', 50))
        cmhc_premium = num(s6, 'C35', 0.05)
    except (KeyError, Exception):
        perm_ltv = 0.95
        perm_dscr = 1.1
        perm_rate = 0.037
        perm_term = 50
        cmhc_premium = 0.05

    # --- Verified final metrics from the Excel's own formulas ---
    # These are the "real" numbers Noor reviewed, not our JS approximations.
    # Read from Exec Summary (Sheet 2) and source sheets (10, 11).
    verified = {}
    try:
        s2 = wb['2. Rev 1B Exec Summary']
        s10 = wb['10. Development Cash Flow']
        s11 = wb['11. 10-Yr Cash Flow IRR']

        v_dev_cost = num(s2, 'G48')         # Total Development Costs
        v_merchant_irr = num(s10, 'G105')   # Merchant Builder IRR
        v_hold_irr = num(s11, 'B67')        # 10-Year Hold IRR
        v_perm_loan = num(s2, 'G59')        # Permanent Loan
        v_annual_debt = num(s2, 'G60')      # Annual Debt Service
        v_ltv = num(s2, 'G57')              # Implied LTV
        v_dscr = num(s2, 'G58')             # DSCR
        v_noi = num(s2, 'H9')              # NOI at Stabilization (base cap)
        v_value = num(s2, 'H11')            # Building Value (base cap)
        v_profit = num(s2, 'H15')           # Profit before taxes (base cap)
        v_merchant_return = num(s2, 'H17')  # Return on Cost (base cap)

        # Cross-check: compare the Exec Summary NOI against the 1A unit data.
        # If the file was generated but never opened in Excel, formula cells
        # still hold the TEMPLATE's cached values (Birchmount), not this project's.
        # A >20% NOI mismatch means the cached values are stale — skip verified.
        expected_annual_rent = sum(u['count'] * u['rent'] * 12 for u in unit_types)  # rough proxy
        noi_plausible = True
        if expected_annual_rent > 0 and v_noi > 0:
            # NOI should be roughly 40-70% of gross rent for typical apartment projects
            noi_ratio = v_noi / expected_annual_rent
            if noi_ratio < 0.1 or noi_ratio > 1.5:
                noi_plausible = False  # stale template values — don't trust

        if noi_plausible:
            if v_dev_cost > 0:
                verified['total_dev_cost'] = v_dev_cost
            if 0 < v_merchant_irr < 5:  # sanity: IRR between 0% and 500%
                verified['merchant_irr'] = v_merchant_irr
            if 0 < v_hold_irr < 5:
                verified['hold_irr'] = v_hold_irr
            if v_perm_loan > 0:
                verified['perm_loan'] = v_perm_loan
            if v_annual_debt > 0:
                verified['annual_debt'] = v_annual_debt
            if 0 < v_ltv < 1:
                verified['ltv'] = v_ltv
            if v_dscr > 0:
                verified['dscr'] = v_dscr
            if v_noi > 0:
                verified['noi'] = v_noi
            if v_value > 0:
                verified['value'] = v_value
            if v_profit != 0:
                verified['profit'] = v_profit
            if 0 < v_merchant_return < 5:
                verified['merchant_return'] = v_merchant_return
    except (KeyError, Exception):
        pass  # Sheet doesn't exist or can't be read — skip verified metrics

    # Building type detection from floor count
    building_type = 'high-rise' if est_floors >= HIGH_RISE_FLOOR_THRESHOLD else 'mid-rise'

    # Detect municipality from address
    muni_name = 'Not selected'
    try:
        municipalities = load_dc_rates()
        for m in municipalities:
            city = m['name'].split(',')[0].split('(')[0].strip().upper()
            if city in address.upper():
                muni_name = m['name']
                break
    except Exception:
        pass

    wb.close()

    # Build the same JSON structure as export_project_json
    project = {
        'project': {
            'name': address.split(',')[0].strip() if address else 'Unknown',
            'address': address,
            'city': ', '.join(address.split(',')[1:]).strip() if ',' in address else '',
            'municipality': muni_name,
            'building_type': building_type,
            'generated': date.today().isoformat(),
            'source': 'reimport',  # flag: this data is from a reviewed Reverse 1B
        },
        'units': {
            'total': total_units,
            'est_storeys': est_floors,
            'types': unit_types,
        },
        'areas': {
            'gfa': gfa,
            'total_rentable_sf': round(total_rentable_sf),
            'amenity_sf': round(amenity_sf),
            'common_area_sf': round(common_area_sf),
            'parking_sf': parking_sf,
        },
        'parking': {
            'underground': {'spaces': pkg_underground, 'fee': pkg_underground_fee},
            'visitor': {'spaces': pkg_visitor, 'fee': pkg_visitor_fee},
            'retail': {'spaces': pkg_retail, 'fee': pkg_retail_fee},
        },
        'storage': {
            'count': storage_count,
            'fee': storage_fee,
        },
        'commercial': {
            'sf': commercial_sf,
            'rate': commercial_rate,
            'vacancy': commercial_vacancy,
        },
        'submetering': {
            'count': total_units,
            'fee': submetering_fee,
        },
        'vacancy_rate': vacancy_rate,
        'cap_rates': {
            'best': cap_best,
            'base': cap_base,
            'worst': cap_worst,
        },
        'opex': {
            'mgmt_fee_pct': mgmt_fee,
            'tax_rate': tax_rate,
            'assessed_value_per_unit': assessed_value,
            'insurance_per_unit': insurance_per_unit,
            'rm_per_unit': rm_per_unit,
            'staffing_per_unit': staffing_per_unit,
            'marketing_per_unit': marketing_per_unit,
            'ga_per_unit': ga_per_unit,
            'utilities_psf': utilities_psf,
            'reserve_pct': reserve_pct,
        },
        'development': {
            'construction_cost_psf': round(construction_cost_psf),
            'soft_cost_pct': 0.30,
            'profit_pct': profit_pct,
            'dc_rates': {
                '1bed': dc_1bed,
                '2bed': dc_2bed,
                '3bed': dc_3bed,
            },
            'dc_total': round(dc_total),
        },
        'financing': {
            'construction_loan_pct': 0.90,
            'construction_loan_rate': 0.0587,
            'perm_loan_ltv': perm_ltv,
            'perm_loan_dscr': perm_dscr,
            'perm_loan_rate': perm_rate,
            'perm_loan_term': perm_term,
            'cmhc_premium': cmhc_premium,
        },
        'schedule': {
            'land_months': land_months,
            'predev_months': predev_months,
            'construction_months': construction_months,
            'leaseup_offset': leaseup_offset,
        },
    }

    # Add verified metrics if we successfully read them from the Excel
    if verified:
        project['verified'] = verified

    return project


# ---------------------------------------------------------------------------
# PROJECT JSON EXPORT — for the presentation tool
# ---------------------------------------------------------------------------

def _blended_construction_rate(cf):
    """Compute blended construction interest rate from debt stack params."""
    cf = cf or {}
    mezz_pct = cf.get('mezz_debt_pct', 0.15)
    mezz_rate = cf.get('mezz_prime_rate', 0.0445) + cf.get('mezz_margin', 0.035)
    bank_pct = cf.get('bank_debt_pct', 0.75)
    bank_rate = cf.get('bank_prime_rate', 0.0445) + cf.get('bank_margin', 0.01)
    total = mezz_pct + bank_pct
    if total == 0:
        return 0
    return (mezz_pct * mezz_rate + bank_pct * bank_rate) / total


def export_project_json(data, output_path, municipality=None, building_type='high-rise',
                        financing_program=None, construction_financing=None):
    """
    Export project data as a JSON file for the presentation/sensitivity tool.
    Contains all inputs needed to calculate revenue, costs, and valuation
    client-side without touching the Excel.
    """
    # Consolidate unit mix same way as populate_template
    if len(data['unit_types']) <= 3:
        consolidated = list(data['unit_types'])
        while len(consolidated) < 3:
            consolidated.append({'label': f"{len(consolidated)+1} Bed", 'sf': 0, 'count': 0, 'rent': 0})
    else:
        consolidated = consolidate_unit_mix(data['unit_types'])

    total_units = data['total_units']
    est_floors = math.ceil(total_units / 12)

    # GFA — from 1A or estimated (populate_template stores back to data)
    gfa = data.get('gfa') or round(data['total_rentable_sf'] / GFA_EFFICIENCY)

    # Amenity SF — use stored value from populate_template, else estimate
    amenity_sf = data.get('amenity_sf') or round(total_units * AMENITY_SF_PER_UNIT, -2)
    # Common area = GFA - rentable SF, matching Excel's utilities formula
    common_area_sf = max(0, gfa - data['total_rentable_sf'])

    # Parking SF
    parking_sf = ((data['parking_underground']['spaces']
                   + data['parking_visitor']['spaces']
                   + data['parking_retail']['spaces']) * PARKING_SF_PER_SPACE)

    # Dev charges total (for permits & approvals)
    dc_total = 0
    dc_rates = {'1bed': 0, '2bed': 0, '3bed': 0}
    if municipality:
        dc_rates = municipality['rates']
        for ut in consolidated:
            label_lower = ut['label'].lower()
            if '1' in label_lower:
                dc_total += ut['count'] * dc_rates['1bed']
            elif '2' in label_lower:
                dc_total += ut['count'] * dc_rates['2bed']
            else:
                dc_total += ut['count'] * dc_rates['3bed']

    # Construction cost per SF — derive from Altus guide baseline
    # The template has construction at ~$453/SF of GFA for high-rise in GTA
    construction_cost_psf = round(65_683_716 / 145_000) if building_type == 'high-rise' else round(45_000_000 / 145_000)

    # Resolve financing program for JSON export
    if financing_program and isinstance(financing_program, dict) and 'label' in financing_program:
        fp = financing_program
        # Find the key by matching the dict
        fp_key = next((k for k, v in FINANCING_PROGRAMS.items() if v is fp and k not in ('cmhc_mli_select', 'cmhc_standard')), 'cmhc_mli_100')
    else:
        fp_key = financing_program if isinstance(financing_program, str) else 'cmhc_mli_100'
        fp = FINANCING_PROGRAMS.get(fp_key, FINANCING_PROGRAMS['cmhc_mli_100'])

    project = {
        'project': {
            'name': data['address'].split(',')[0].strip(),
            'address': data['address'],
            'city': ', '.join(data['address'].split(',')[1:]).strip() if ',' in data['address'] else '',
            'municipality': municipality['name'] if municipality else 'Not selected',
            'building_type': building_type,
            'generated': date.today().isoformat(),
        },
        'units': {
            'total': total_units,
            'est_storeys': est_floors,
            'types': [
                {'label': u['label'], 'count': u['count'], 'sf': u['sf'], 'rent': u['rent']}
                for u in consolidated if u['count'] > 0
            ],
        },
        'areas': {
            'gfa': gfa,
            'total_rentable_sf': round(data['total_rentable_sf']),
            'amenity_sf': amenity_sf,
            'common_area_sf': common_area_sf,
            'parking_sf': parking_sf,
        },
        'parking': {
            'underground': {'spaces': data['parking_underground']['spaces'],
                            'fee': data['parking_underground']['fee']},
            'visitor': {'spaces': data['parking_visitor']['spaces'],
                        'fee': data['parking_visitor']['fee']},
            'retail': {'spaces': data['parking_retail']['spaces'],
                       'fee': data['parking_retail']['fee']},
        },
        'storage': {
            'count': data['storage']['count'],
            'fee': data['storage']['fee'],
        },
        'commercial': {
            'sf': data['commercial']['sf'],
            'rate': data['commercial']['rate'],
            'vacancy': data['commercial_vacancy'],
        },
        'submetering': data.get('submetering', 0),
        'vacancy_rate': data['vacancy_rate'],
        'cap_rates': {
            'best': data['cap_rates'][0] if len(data['cap_rates']) > 0 else 0.0425,
            'base': data['cap_rates'][1] if len(data['cap_rates']) > 1 else 0.045,
            'worst': data['cap_rates'][2] if len(data['cap_rates']) > 2 else 0.0475,
        },
        'opex': {
            'mgmt_fee_pct': data.get('mgmt_fee_pct', 0.0425),
            'tax_rate': data.get('tax_rate', 0),
            'assessed_value_per_unit': data.get('assessed_value', 0),
            # Expenses live in data['expenses'] dict from parser; fall back to defaults
            'insurance_per_unit': data.get('expenses', {}).get('insurance', DEFAULT_INSURANCE_PER_UNIT),
            'rm_per_unit': data.get('expenses', {}).get('rm', DEFAULT_RM_PER_UNIT),
            'staffing_per_unit': data.get('expenses', {}).get('staffing', DEFAULT_STAFFING_PER_UNIT),
            'marketing_per_unit': data.get('expenses', {}).get('marketing', DEFAULT_MARKETING_PER_UNIT),
            'ga_per_unit': data.get('expenses', {}).get('ga', DEFAULT_GA_PER_UNIT),
            'utilities_psf': data.get('utilities_psf', DEFAULT_UTILITIES_PSF),
            'reserve_pct': data.get('reserve_pct', DEFAULT_RESERVE_PCT),
        },
        'development': {
            'construction_cost_psf': construction_cost_psf,
            'soft_cost_pct': 0.30,
            'profit_pct': 0.08,
            'dc_rates': dc_rates,
            'dc_total': round(dc_total),
        },
        'financing': {
            # Construction financing (Sheet 5 rows 67-71)
            'construction_loan_pct': (construction_financing or {}).get('mezz_debt_pct', 0.15) + (construction_financing or {}).get('bank_debt_pct', 0.75),
            'construction_loan_rate': _blended_construction_rate(construction_financing),
            # Permanent take-out loan (Sheet 6 rows 27-35) — from selected program
            'program_key': fp_key,
            'program_label': fp['label'],
            'perm_loan_ltv': fp['max_ltv'],
            'perm_loan_dscr': fp['min_dscr'],
            'perm_loan_rate': fp['interest_rate'],
            'perm_loan_term': fp['amortization'],
            'cmhc_premium': fp['cmhc_premium'],
        },
        'schedule': {
            'land_months': 0,
            'predev_months': 12,
            'construction_months': 18,
            'leaseup_offset': -3,
        },
    }

    with open(output_path, 'w') as f:
        json.dump(project, f, indent=2)

    return project


# ---------------------------------------------------------------------------
# EXTRACT VERIFIED METRICS — from actual Excel formula evaluation
# ---------------------------------------------------------------------------

def _extract_excel_metrics(xlsx_path):
    """
    Extract verified metrics by recalculating the xlsx with LibreOffice
    headless mode, then reading the cached formula results with openpyxl.

    This gives us EVERY value Excel would show — no estimation, no
    proportional splits, no formulas library limitations. LibreOffice
    evaluates the full Altus CHOOSE chain, TODAY()-based escalation,
    construction interest draw schedules, and all other formulas that
    the Python 'formulas' library can't handle.

    Falls back to the 'formulas' library if LibreOffice is unavailable.
    """
    # Try LibreOffice recalculation first — gives ALL values exactly
    recalc_result = _extract_via_libreoffice(xlsx_path)
    if recalc_result:
        return recalc_result

    # No LibreOffice available (e.g. Railway). Skip the formulas library
    # entirely — it tries to resolve external workbook refs from Noor's
    # local Windows paths, consuming all available memory and crashing
    # the server. Python metrics are accurate to within 0.06pt on IRRs
    # and exact on everything else. Not worth the crash risk.
    return None


def _extract_via_libreoffice(xlsx_path):
    """Recalculate with LibreOffice headless and read all cached values."""
    import subprocess
    import tempfile
    import shutil

    # Find LibreOffice
    soffice = None
    for path in ['/Applications/LibreOffice.app/Contents/MacOS/soffice',
                 '/opt/homebrew/bin/soffice', '/usr/bin/soffice',
                 '/usr/local/bin/soffice']:
        if os.path.exists(path):
            soffice = path
            break
    if not soffice:
        return None

    try:
        # Copy to temp dir for recalculation (don't modify the output file)
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_xlsx = os.path.join(tmpdir, os.path.basename(xlsx_path))
            shutil.copy2(xlsx_path, tmp_xlsx)

            # Recalculate and re-export as xlsx
            proc = subprocess.run(
                [soffice, '--headless', '--calc', '--convert-to', 'xlsx',
                 '--outdir', tmpdir, tmp_xlsx],
                capture_output=True, timeout=30
            )
            if proc.returncode != 0:
                return None

            # Read recalculated values
            import openpyxl
            wb = openpyxl.load_workbook(tmp_xlsx, data_only=True)

            def get_cell(sheet_name, cell_ref):
                """Read a cell's cached value after recalculation."""
                try:
                    ws = wb[sheet_name]
                    val = ws[cell_ref].value
                    if val is not None and isinstance(val, (int, float)):
                        return float(val)
                except (KeyError, TypeError):
                    pass
                return None

            # Find sheets by prefix — names vary slightly across template versions
            sheet_names = wb.sheetnames
            S2 = next((s for s in sheet_names if s.startswith('2.')), None)
            S5 = next((s for s in sheet_names if s.startswith('5.')), None)
            S9 = next((s for s in sheet_names if s.startswith('9.')), None)

            if not all([S2, S5, S9]):
                wb.close()
                return None

            noi = get_cell(S2, 'H9')
            value = get_cell(S2, 'H11')
            if not noi or not value:
                wb.close()
                return None

            result = {'noi': noi, 'value': value}

            # Sheet 2 — financing metrics (including construction loan/equity for IRR)
            for key, cell in [('dscr', 'G58'), ('ltv', 'G57'),
                              ('perm_loan', 'G59'), ('annual_debt', 'G60'),
                              ('constr_loan', 'G52'), ('constr_equity', 'G53')]:
                v = get_cell(S2, cell)
                if v:
                    result[key] = v

            # Sheet 5 — TDC
            tdc = get_cell(S5, 'F38')
            if tdc and tdc > 0:
                result['total_dev_cost'] = tdc
            budget = get_cell(S5, 'F36')
            if budget and budget > 0:
                result['total_project_budget'] = budget

            # Sheet 9 — ALL dev cost line items (exact after recalculation)
            dev_items = {
                'dev_land':         'H16',   # Total land costs
                'dev_construction': 'H22',   # Total construction costs
                'dev_prof_fees':    'H41',   # Total professional fees
                'dev_mgmt':         'H46',   # Total dev management fee
                'dev_permits':      'H59',   # Total permits & approvals
                'dev_marketing':    'H66',   # Total marketing & leasing
                'dev_financing':    'H80',   # Total financing & interest
                'dev_lease_up':     'H87',   # Total lease-up income
            }
            for key, cell in dev_items.items():
                v = get_cell(S9, cell)
                if v is not None:
                    result[key] = v

            # TDC from Sheet 9 as cross-check
            tdc_s9 = get_cell(S9, 'H89')
            if tdc_s9:
                result['total_dev_cost_s9'] = tdc_s9

            # IRRs directly from Excel (exact, no approximation needed)
            S10 = next((s for s in sheet_names if s.startswith('10.')), None)
            S11 = next((s for s in sheet_names if s.startswith('11.')), None)
            if S10:
                mirr = get_cell(S10, 'G105')
                if mirr and 0 < mirr < 5:
                    result['merchant_irr_excel'] = mirr
            if S11:
                hirr = get_cell(S11, 'B67')
                if hirr and 0 < hirr < 5:
                    result['hold_irr_excel'] = hirr

            wb.close()
            return result

    except Exception as e:
        print(f"  [libreoffice] Warning: {e}")
        return None


def _extract_via_formulas_library(xlsx_path):
    """Fallback: use Python formulas library (partial coverage)."""
    try:
        import formulas
        import numpy as np
        import warnings
        import io
        import sys as _sys
        import signal
        import threading
        warnings.filterwarnings('ignore')

        # Timeout guard: the formulas library can hang on Railway trying to
        # resolve external workbook refs (Noor's local Windows paths).
        # Cap at 30 seconds to prevent gunicorn worker OOM/SIGKILL.
        result_holder = [None]
        error_holder = [None]

        def _run_formulas():
            try:
                old_stderr = _sys.stderr
                _sys.stderr = io.StringIO()
                try:
                    xl = formulas.ExcelModel().loads(xlsx_path).finish()
                    result_holder[0] = xl.calculate()
                finally:
                    _sys.stderr = old_stderr
            except Exception as e:
                error_holder[0] = e

        t = threading.Thread(target=_run_formulas, daemon=True)
        t.start()
        t.join(timeout=30)
        if t.is_alive():
            print(f"  [formulas] Warning: timed out after 30s, skipping formula evaluation")
            return None
        if error_holder[0]:
            raise error_holder[0]
        sol = result_holder[0]
        if sol is None:
            return None

        def get_val(sheet_part, cell):
            cell_upper = cell.upper()
            sheet_upper = sheet_part.upper()
            for k, val in sol.items():
                ks = str(k).upper()
                if sheet_upper not in ks:
                    continue
                if ks.endswith("!" + cell_upper) or ks.endswith("'!" + cell_upper):
                    try:
                        if hasattr(val, 'value'):
                            v = float(np.array(val.value).flatten()[0])
                        else:
                            v = float(np.array(val).flatten()[0])
                        if math.isnan(v) or math.isinf(v):
                            return None
                        return v
                    except:
                        return None
            return None

        S2 = "EXEC SUMMARY"
        S5 = "KEY ASSUMPTION"

        noi = get_val(S2, "H9")
        value = get_val(S2, "H11")
        if not noi or not value:
            return None

        result = {'noi': noi, 'value': value}

        for key, cell in [('dscr', 'G58'), ('ltv', 'G57'),
                          ('perm_loan', 'G59'), ('annual_debt', 'G60')]:
            v = get_val(S2, cell)
            if v:
                result[key] = v

        tdc = get_val(S5, "F38")
        if tdc and tdc > 0:
            result['total_dev_cost'] = tdc
        budget = get_val(S5, "F36")
        if budget and budget > 0:
            result['total_project_budget'] = budget

        return result

    except ImportError:
        return None
    except Exception as e:
        print(f"  [formulas] Warning: {e}")
        return None


# ---------------------------------------------------------------------------
# CALCULATE VERIFIED METRICS
# ---------------------------------------------------------------------------

def calculate_verified_metrics(project, xlsx_path=None):
    """
    Build the most accurate metrics possible by combining Excel formula
    evaluation with precise Python derivation.

    Strategy — extract exact values from Excel's formula chain, cascade the rest:

    EXACT from Excel (Sheet 2):
      NOI, Value, DSCR, LTV, Perm Loan, Annual Debt

    EXACT from Excel (Sheet 5):
      Total Development Cost (F38) — the single most important number

    DERIVED from exact TDC + exact Value:
      Profit, Merchant Return, Merchant IRR, Hold IRR

    The only remaining approximation is IRR timing (monthly cash flow model),
    but the dollar amounts feeding it are all exact from Excel.
    """
    py_metrics = _calculate_python_metrics(project)

    if xlsx_path:
        excel = _extract_excel_metrics(xlsx_path)
        if excel and py_metrics:
            # Override revenue/financing metrics with exact Excel values
            py_metrics['noi'] = excel['noi']
            py_metrics['value'] = excel['value']
            for k in ('dscr', 'ltv', 'perm_loan', 'annual_debt'):
                if k in excel:
                    py_metrics[k] = excel[k]

            exact_keys = {'noi', 'value', 'dscr', 'ltv', 'perm_loan', 'annual_debt'}

            P = project
            fin = P.get('financing', {})
            sched = P.get('schedule', {})
            units_total = P['units']['total']

            # ── TDC: prefer Sheet 9 H89 (bottom-up), fall back to Sheet 5 F38 ──
            # Sheet 9 H89 sums all 8 cost categories and is the authoritative TDC.
            # Sheet 5 F38 is a top-down formula that diverges for non-template projects.
            if 'dev_construction' in excel:
                # Dev breakdown available — use Sheet 9 H89 as TDC
                breakdown = {
                    'land':         excel.get('dev_land', 0),
                    'construction': excel.get('dev_construction', 0),
                    'prof_fees':    excel.get('dev_prof_fees', 0),
                    'dev_mgmt':     excel.get('dev_mgmt', 0),
                    'permits':      excel.get('dev_permits', 0),
                    'marketing':    excel.get('dev_marketing', 0),
                    'financing':    excel.get('dev_financing', 0),
                    'lease_up':     excel.get('dev_lease_up', 0),
                }
                py_metrics['dev_breakdown'] = breakdown
                exact_keys.add('dev_breakdown')
                s9_total = excel.get('total_dev_cost_s9')
                if s9_total and s9_total > 0:
                    total_dev_cost = s9_total
                    py_metrics['total_dev_cost'] = s9_total
                    exact_keys.add('total_dev_cost')
                elif 'total_dev_cost' in excel:
                    total_dev_cost = excel['total_dev_cost']
                    py_metrics['total_dev_cost'] = total_dev_cost
                    exact_keys.add('total_dev_cost')
                else:
                    total_dev_cost = py_metrics.get('total_dev_cost', 0)
            elif 'total_dev_cost' in excel:
                total_dev_cost = excel['total_dev_cost']
                py_metrics['total_dev_cost'] = total_dev_cost
                exact_keys.add('total_dev_cost')
            else:
                total_dev_cost = py_metrics.get('total_dev_cost', 0)

            # ── Cascade: profit, returns, IRRs from exact TDC + exact value ──
            selling_cost_pct = 0.01
            sales_proceeds = excel['value'] * (1 - selling_cost_pct)

            if total_dev_cost > 0:
                py_metrics['profit'] = sales_proceeds - total_dev_cost
                py_metrics['merchant_return'] = py_metrics['profit'] / total_dev_cost
                # Profit and return are pure arithmetic from exact TDC + exact Value
                if 'total_dev_cost' in exact_keys:
                    exact_keys.update(['profit', 'merchant_return'])

                # Use actual construction loan/equity from Excel if available
                # (the loan is sized by formula, NOT always 90% of TDC)
                if 'constr_loan' in excel and 'constr_equity' in excel:
                    constr_debt = excel['constr_loan']
                    constr_equity = excel['constr_equity']
                else:
                    constr_loan_pct = fin.get('construction_loan_pct', 0.90)
                    constr_equity = total_dev_cost * (1 - constr_loan_pct)
                    constr_debt = total_dev_cost * constr_loan_pct
                predev = sched.get('predev_months', 12)
                construction_mo = sched.get('construction_months', 18)
                lease_up_months = math.ceil(units_total / 15)
                # +1 for stabilization month (matches Sheet 10: sale at month N+1)
                total_dev_months = predev + construction_mo + lease_up_months + 1

                # Merchant IRR — CAGR over total_dev_months periods
                # Sheet 10: equity at month 0, sale at month N, all intermediate CFs = 0
                merchant_proceeds = sales_proceeds - constr_debt
                if constr_equity > 0 and merchant_proceeds > 0 and total_dev_months > 0:
                    monthly_r = (merchant_proceeds / constr_equity) ** (1 / total_dev_months) - 1
                    merchant_irr = (1 + monthly_r) ** 12 - 1
                    if 0 < merchant_irr < 5:
                        py_metrics['merchant_irr'] = merchant_irr
                        if 'total_dev_cost' in exact_keys:
                            exact_keys.add('merchant_irr')

                # Hold IRR — 10-year annual cash flow model (matches Sheet 11)
                perm_rate = fin.get('perm_loan_rate', 0.037)
                perm_amort = fin.get('perm_loan_term', 50)
                perm_loan = py_metrics.get('perm_loan', 0)
                annual_debt = py_metrics.get('annual_debt', 0)
                cmhc_premium = perm_loan * fin.get('cmhc_premium', 0.05)
                noi_stab = excel['noi']
                cap_base = P['cap_rates']['base']

                # dev_years = calendar year when stabilization occurs
                # stab_month = total months from start to stabilization (1-indexed)
                stab_month = predev + construction_mo + lease_up_months + 2
                dev_years = math.ceil(stab_month / 12)
                # Stabilized months in the refi year (remaining months after stabilization)
                partial_months = dev_years * 12 - stab_month

                refi_cash_out = perm_loan - constr_debt - cmhc_premium
                partial_year_income = (noi_stab - annual_debt) * (partial_months / 12)

                if constr_equity > 0:
                    hold_years = 10
                    # Year 1: equity investment
                    cf = [-constr_equity]
                    # Years 2 through dev_years-1: zero (development)
                    for y in range(2, dev_years):
                        cf.append(0)
                    # Year dev_years: CMHC refi + partial year income
                    cf.append(refi_cash_out + partial_year_income)
                    # Years dev_years+1 through hold_years-1: full operating income
                    for y in range(dev_years + 1, hold_years):
                        years_from_stab = y - dev_years
                        cf.append(noi_stab * (1.02 ** years_from_stab) - annual_debt)
                    # Year hold_years (10): operating income + exit sale
                    exit_years_from_stab = hold_years - dev_years
                    exit_noi = noi_stab * (1.02 ** exit_years_from_stab)
                    exit_value = exit_noi / cap_base if cap_base > 0 else 0
                    monthly_rate = perm_rate / 12
                    n_months = perm_amort * 12
                    monthly_pmt = (monthly_rate * perm_loan) / (1 - (1 + monthly_rate) ** (-n_months)) if monthly_rate > 0 else 0
                    # Exit debt = balance at START of final year (matches Sheet 11 Row 34)
                    exit_debt_months = partial_months + (hold_years - dev_years - 1) * 12
                    bal_growth = (1 + monthly_rate) ** exit_debt_months
                    if monthly_rate > 0:
                        exit_debt_balance = perm_loan * bal_growth - monthly_pmt * (bal_growth - 1) / monthly_rate
                    else:
                        exit_debt_balance = perm_loan * (1 - exit_debt_months / (perm_amort * 12))
                    exit_proceeds = exit_value * (1 - selling_cost_pct) - max(0, exit_debt_balance)
                    cf.append((noi_stab * (1.02 ** exit_years_from_stab) - annual_debt) + exit_proceeds)

                    guess = _solve_irr(cf)
                    if 0 < guess < 5:
                        py_metrics['hold_irr'] = guess
                        if 'total_dev_cost' in exact_keys:
                            exact_keys.add('hold_irr')

            # Override IRRs with exact Excel values when available
            # (our JS approximation diverges for high-return projects)
            if 'merchant_irr_excel' in excel:
                py_metrics['merchant_irr'] = excel['merchant_irr_excel']
                exact_keys.add('merchant_irr')
            if 'hold_irr_excel' in excel:
                py_metrics['hold_irr'] = excel['hold_irr_excel']
                exact_keys.add('hold_irr')

            py_metrics['exact_keys'] = list(exact_keys)
            return py_metrics
        elif excel:
            excel['exact_keys'] = list(excel.keys())
            return excel

    return py_metrics


def _calculate_python_metrics(project):
    """
    Calculate all key financial metrics from a project JSON dict.
    Mirrors the JS calculate() function in presentation.html exactly,
    so baseline numbers match with zero discrepancy.

    Returns a verified dict with the same keys as import_reverse_1b()'s
    verified output, or None on failure.
    """
    try:
        P = project
        units_total = P['units']['total']
        if units_total <= 0:
            return None

        unit_types = P['units']['types']
        total_rentable_sf = sum(u['count'] * u['sf'] for u in unit_types)

        # Revenue
        residential_rent = sum(u['count'] * u['rent'] * 12 for u in unit_types)
        parking_revenue = (
            P['parking']['underground']['spaces'] * P['parking']['underground']['fee']
            + P['parking']['visitor']['spaces'] * P['parking']['visitor']['fee']
            + P['parking']['retail']['spaces'] * P['parking']['retail']['fee']
        ) * 12
        storage_revenue = P['storage']['count'] * P['storage']['fee'] * 12
        sub = P.get('submetering', 0)
        if isinstance(sub, dict):
            submetering = (sub.get('count', 0) or 0) * (sub.get('fee', 0) or 0) * 12
        else:
            submetering = sub or 0
        commercial_revenue = P['commercial']['sf'] * P['commercial']['rate']
        commercial_vacancy_rate = P['commercial'].get('vacancy', 0)

        # Match Excel: residential vacancy applies to rent+parking+storage+submetering
        # Commercial has its own separate vacancy rate
        residential_subtotal = residential_rent + parking_revenue + storage_revenue + submetering
        vacancy_rate = P['vacancy_rate']
        residential_vacancy = residential_subtotal * vacancy_rate
        commercial_vacancy = commercial_revenue * commercial_vacancy_rate
        egi = residential_subtotal - residential_vacancy + commercial_revenue - commercial_vacancy

        # OpEx
        opex = P['opex']
        mgmt_fee = egi * opex['mgmt_fee_pct']
        property_tax = opex['tax_rate'] * opex['assessed_value_per_unit'] * units_total
        insurance = opex['insurance_per_unit'] * units_total
        rm = opex['rm_per_unit'] * units_total
        staffing = opex['staffing_per_unit'] * units_total
        marketing = opex['marketing_per_unit'] * units_total
        ga = opex['ga_per_unit'] * units_total
        common_area = max(0, P['areas']['gfa'] - P['areas']['total_rentable_sf'])
        utilities = opex['utilities_psf'] * common_area
        reserve = opex['reserve_pct'] * egi
        total_opex = mgmt_fee + property_tax + insurance + rm + staffing + marketing + ga + utilities + reserve

        noi = egi - total_opex

        # Development timeline
        sched = P.get('schedule', {})
        predev = sched.get('predev_months', 12)
        construction = sched.get('construction_months', 18)
        lease_up_months = math.ceil(units_total / 15)
        # Stabilization month (1-indexed) and inflation years
        # Inflation = 1.02^(refi_year - 1) where refi_year = calendar year of stabilization
        stab_month = predev + construction + lease_up_months + 2
        dev_years = math.ceil(stab_month / 12)
        noi_stabilized = noi * (1.02 ** (dev_years - 1))

        # Valuation
        cap_base = P['cap_rates']['base']
        if cap_base <= 0:
            return None
        value_base = noi_stabilized / cap_base
        selling_cost_pct = 0.01
        sales_proceeds = value_base * (1 - selling_cost_pct)

        # ── Development costs (mirrors Sheet 9 line-by-line) ──
        construction_psf = P['development']['construction_cost_psf']
        gfa = P['areas']['gfa']
        parking_sf = P['areas'].get('parking_sf', 0)

        # 1. LAND — backed out from stabilized value
        #    Sheet 5: land = value × (profit% + 2.5%)
        profit_pct = P['development'].get('profit_pct', 0.08)
        land_cost = value_base * (profit_pct + 0.025)
        land_closing = land_cost * 0.05
        total_land = land_cost + land_closing

        # 2. CONSTRUCTION
        hard_costs = construction_psf * gfa
        construction_contingency = hard_costs * 0.02
        total_construction = hard_costs + construction_contingency

        # 3. PROFESSIONAL FEES
        architect = hard_costs * 0.03
        fee_contingency = architect * 0.05
        total_prof_fees = architect + fee_contingency

        # 4. DEVELOPMENT MANAGEMENT — 2.5% of (HC + contingency + fees)
        dev_mgmt = (hard_costs + construction_contingency + architect + fee_contingency) * 0.025

        # 5. PERMITS & APPROVALS
        dc_total = P['development']['dc_total']
        submeter_credit = 600 * units_total  # $600/unit credit
        permit_contingency = dc_total * 0.05
        total_permits = dc_total - submeter_credit + permit_contingency

        # 6. MARKETING & LEASING — 1.5 months of (rent + parking) revenue
        monthly_revenue = (residential_rent + parking_revenue) / 12
        commissions = monthly_revenue * 1.5
        mktg_cost = commissions  # marketing = same as commissions
        total_marketing = commissions + mktg_cost

        # Sum pre-financing costs
        pre_financing = (total_land + total_construction + total_prof_fees
                         + dev_mgmt + total_permits + total_marketing)

        # 7. FINANCING — loan fees + interest on construction draws
        #    Total debt = 90% of total dev cost (circular ref — iterate)
        #    Loan fees = 1% of debt, financing contingency = 0.5% of debt
        #    Construction interest from monthly draw schedule
        fin = P['financing']
        constr_loan_pct = fin.get('construction_loan_pct', 0.90)
        constr_rate = fin.get('construction_loan_rate', 0.0587)

        # Monthly draw schedule for construction interest
        # Pre-dev: costs drawn evenly, construction: drawn evenly
        predev_cost = total_land + total_permits * 0.5  # land + half permits during predev
        construction_draw = total_construction + total_prof_fees + dev_mgmt + total_permits * 0.5 + total_marketing
        monthly_interest = 0
        cumulative = 0
        for m in range(predev + construction):
            if m < predev:
                cumulative += predev_cost / predev if predev > 0 else 0
            else:
                cumulative += construction_draw / construction if construction > 0 else 0
            monthly_interest += cumulative * (constr_rate / 12)

        total_debt_est = pre_financing * constr_loan_pct
        loan_fees = total_debt_est * 0.01
        financing_contingency = total_debt_est * 0.005
        total_financing = loan_fees + financing_contingency + monthly_interest

        # 8. LEASE-UP INCOME (negative cost — reduces TDC)
        monthly_gross = residential_rent / 12 + parking_revenue / 12 + storage_revenue / 12
        lease_up_income = 0
        lease_up_expenses = 0
        monthly_opex = total_opex / 12
        for m in range(lease_up_months):
            occ = min((m + 1) * 15, units_total) / units_total
            lease_up_income += monthly_gross * occ
            lease_up_expenses += monthly_opex * occ
        # Vacancy during lease-up
        lease_up_vacancy = lease_up_income * vacancy_rate
        net_lease_up = lease_up_income - lease_up_vacancy - lease_up_expenses

        total_dev_cost = pre_financing + total_financing - net_lease_up
        dev_profit = sales_proceeds - total_dev_cost
        merchant_return = dev_profit / total_dev_cost if total_dev_cost > 0 else 0

        # CMHC permanent loan (with defaults for older JSON formats)
        perm_ltv = fin.get('perm_loan_ltv', 0.95)
        perm_dscr = fin.get('perm_loan_dscr', 1.1)
        perm_rate = fin.get('perm_loan_rate', 0.037)
        perm_amort = fin.get('perm_loan_term', 50)
        cmhc_premium_rate = fin.get('cmhc_premium', 0.05)

        ltv_loan = perm_ltv * value_base
        allowed_annual_pmt = noi_stabilized / perm_dscr
        pv_factor = (1 - (1 + perm_rate) ** (-perm_amort)) / perm_rate if perm_rate > 0 else 0
        dscr_loan = allowed_annual_pmt * pv_factor
        perm_loan = min(ltv_loan, dscr_loan)

        monthly_rate = perm_rate / 12
        n_months = perm_amort * 12
        if monthly_rate > 0:
            monthly_pmt = (monthly_rate * perm_loan) / (1 - (1 + monthly_rate) ** (-n_months))
        else:
            monthly_pmt = 0
        annual_debt = monthly_pmt * 12
        dscr = noi_stabilized / annual_debt if annual_debt > 0 else 0
        cmhc_premium = perm_loan * cmhc_premium_rate
        implied_ltv = perm_loan / value_base if value_base > 0 else 0

        # Construction financing
        constr_equity = total_dev_cost * (1 - constr_loan_pct)
        constr_debt = total_dev_cost * constr_loan_pct

        # Merchant IRR — CAGR over total_dev_months periods
        # +1 for stabilization month (matches Sheet 10: sale at month N+1)
        merchant_proceeds = sales_proceeds - constr_debt
        total_dev_months = predev + construction + lease_up_months + 1
        if constr_equity > 0 and merchant_proceeds > 0 and total_dev_months > 0:
            monthly_r = (merchant_proceeds / constr_equity) ** (1 / total_dev_months) - 1
            merchant_irr = (1 + monthly_r) ** 12 - 1
        else:
            merchant_irr = 0

        # Hold IRR — 10-year annual cash flow model (matches Sheet 11)
        hold_years = 10
        # dev_years uses full lease-up duration, not just the offset
        stab_month = predev + construction + lease_up_months + 2
        dev_years = math.ceil(stab_month / 12)
        partial_months = dev_years * 12 - stab_month

        refi_cash_out = perm_loan - constr_debt - cmhc_premium
        partial_year_income = (noi_stabilized - annual_debt) * (partial_months / 12)

        hold_irr = 0
        if constr_equity > 0:
            cf = [-constr_equity]
            for y in range(2, dev_years):
                cf.append(0)
            cf.append(refi_cash_out + partial_year_income)
            for y in range(dev_years + 1, hold_years):
                years_from_stab = y - dev_years
                cf.append(noi_stabilized * (1.02 ** years_from_stab) - annual_debt)
            # Exit year (year hold_years)
            exit_years_from_stab = hold_years - dev_years
            exit_noi = noi_stabilized * (1.02 ** exit_years_from_stab)
            exit_value = exit_noi / cap_base if cap_base > 0 else 0
            # Exit debt = balance at START of final year (matches Sheet 11 Row 34)
            exit_debt_months = partial_months + (hold_years - dev_years - 1) * 12
            bal_growth = (1 + monthly_rate) ** exit_debt_months
            if monthly_rate > 0:
                exit_debt_balance = perm_loan * bal_growth - monthly_pmt * (bal_growth - 1) / monthly_rate
            else:
                exit_debt_balance = perm_loan * (1 - exit_debt_months / (perm_amort * 12))
            exit_proceeds = exit_value * (1 - selling_cost_pct) - max(0, exit_debt_balance)
            cf.append((noi_stabilized * (1.02 ** exit_years_from_stab) - annual_debt) + exit_proceeds)

            hold_irr = _solve_irr(cf)

        # Build verified dict (same keys as import_reverse_1b)
        verified = {}
        if total_dev_cost > 0:
            verified['total_dev_cost'] = total_dev_cost
        if 0 < merchant_irr < 5:
            verified['merchant_irr'] = merchant_irr
        if 0 < hold_irr < 5:
            verified['hold_irr'] = hold_irr
        if perm_loan > 0:
            verified['perm_loan'] = perm_loan
        if annual_debt > 0:
            verified['annual_debt'] = annual_debt
        if 0 < implied_ltv < 1:
            verified['ltv'] = implied_ltv
        if dscr > 0:
            verified['dscr'] = dscr
        if noi_stabilized > 0:
            verified['noi'] = noi_stabilized
        if value_base > 0:
            verified['value'] = value_base
        if dev_profit != 0:
            verified['profit'] = dev_profit
        if 0 < merchant_return < 5:
            verified['merchant_return'] = merchant_return

        return verified

    except (KeyError, ZeroDivisionError, Exception):
        return None


def recalculate_and_extract(xlsx_path):
    """
    Legacy function signature — kept for backward compatibility.
    Now returns None since LibreOffice headless doesn't reliably
    recalculate formulas. Use calculate_verified_metrics() instead.
    """
    return None


# ---------------------------------------------------------------------------
# REIMPORT DIFF — compare original generation vs Noor's reviewed version
# ---------------------------------------------------------------------------

def diff_projects(original, reimported):
    """
    Compare two project JSON dicts and return human-readable changes.
    Useful for understanding what Noor adjusted during review.
    """
    changes = []
    metrics = {}

    def _cmp(path, orig_val, new_val, label, fmt=',.0f', threshold=0.001):
        """Compare two numeric values and record the change."""
        if orig_val is None or new_val is None:
            return
        try:
            o, n = float(orig_val), float(new_val)
        except (ValueError, TypeError):
            return
        if abs(o) < 0.0001 and abs(n) < 0.0001:
            return
        if abs(o) > 0.0001:
            pct = (n - o) / abs(o)
        else:
            pct = 1.0 if n != 0 else 0
        if abs(pct) > threshold:
            direction = '+' if n > o else ''
            changes.append(f"{label}: {format(o, fmt)} → {format(n, fmt)} ({direction}{pct:.1%})")
            metrics[path] = {'original': o, 'new': n, 'pct_change': round(pct, 4)}

    # OpEx comparisons
    for key, label in [
        ('insurance_per_unit', 'Insurance/unit'),
        ('rm_per_unit', 'R&M/unit'),
        ('staffing_per_unit', 'Staffing/unit'),
        ('marketing_per_unit', 'Marketing/unit'),
        ('ga_per_unit', 'G&A/unit'),
        ('utilities_psf', 'Utilities $/PSF'),
        ('tax_rate', 'Tax Rate'),
        ('assessed_value_per_unit', 'Assessed Value/Unit'),
        ('mgmt_fee_pct', 'Mgmt Fee %'),
        ('reserve_pct', 'Reserve %'),
    ]:
        o_val = original.get('opex', {}).get(key)
        n_val = reimported.get('opex', {}).get(key)
        fmt = '.4f' if 'pct' in key or 'rate' in key else ',.0f'
        _cmp(f'opex.{key}', o_val, n_val, label, fmt=fmt)

    # Cap rates
    for tier in ('best', 'base', 'worst'):
        o_val = original.get('cap_rates', {}).get(tier)
        n_val = reimported.get('cap_rates', {}).get(tier)
        _cmp(f'cap_rates.{tier}', o_val, n_val, f'Cap Rate ({tier})', fmt='.4f')

    # Vacancy
    _cmp('vacancy_rate', original.get('vacancy_rate'), reimported.get('vacancy_rate'),
         'Vacancy Rate', fmt='.4f')

    # DC rates
    for bed in ('1bed', '2bed', '3bed'):
        o_val = original.get('development', {}).get('dc_rates', {}).get(bed)
        n_val = reimported.get('development', {}).get('dc_rates', {}).get(bed)
        _cmp(f'dc_rates.{bed}', o_val, n_val, f'DC Rate ({bed})', fmt=',.0f')

    # Verified metrics (if both have them)
    o_ver = original.get('verified', {})
    n_ver = reimported.get('verified', {})
    if o_ver and n_ver:
        for key, label in [
            ('noi', 'NOI'),
            ('total_dev_cost', 'Total Dev Cost'),
            ('merchant_irr', 'Merchant IRR'),
            ('hold_irr', 'Hold IRR'),
            ('perm_loan', 'Permanent Loan'),
            ('dscr', 'DSCR'),
            ('ltv', 'LTV'),
        ]:
            fmt = '.4f' if key in ('merchant_irr', 'hold_irr', 'dscr', 'ltv') else ',.0f'
            _cmp(f'verified.{key}', o_ver.get(key), n_ver.get(key), label, fmt=fmt)

    return {'changes': changes, 'metrics': metrics}


# ---------------------------------------------------------------------------
# MUNICIPALITY GAP TRACKING — log when a municipality isn't in our data
# ---------------------------------------------------------------------------

def _log_municipality_gap(address, attempted, output_dir):
    """Append a missing-municipality entry to the gap log."""
    gap_path = os.path.join(output_dir, 'municipality_gaps.json')
    try:
        if os.path.exists(gap_path):
            with open(gap_path) as f:
                gaps = json.load(f)
        else:
            gaps = []
        gaps.append({
            'date': date.today().isoformat(),
            'address': address,
            'attempted': attempted,
        })
        with open(gap_path, 'w') as f:
            json.dump(gaps, f, indent=2)
    except Exception:
        pass  # non-critical — don't break generation


# ---------------------------------------------------------------------------
# CLI ENTRY POINT
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python populate_reverse_1b.py <path_to_1a_proforma>")
        print("Example: python populate_reverse_1b.py reference/1A_Birchmount_2240.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]
    if not os.path.exists(input_path):
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    if not os.path.exists(TEMPLATE_PATH):
        print(f"Error: Template not found: {TEMPLATE_PATH}")
        sys.exit(1)

    # Create output directory
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Parse the 1A
    print(f"Parsing 1A proforma: {input_path}")
    data = parse_1a(input_path)

    # Print parsed data summary
    print(f"\nProject: {data['address']}")
    print(f"Unit types found: {len(data['unit_types'])}")
    for u in data['unit_types']:
        print(f"  {u['label']}: {u['count']} units, {u['sf']} SF, ${u['rent']}/mo")
    print(f"Total units: {data['total_units']}")
    print(f"Total rentable SF: {data['total_rentable_sf']:,.0f}")
    print(f"Parking: {data['parking_underground']['spaces']} underground, "
          f"{data['parking_visitor']['spaces']} visitor, {data['parking_retail']['spaces']} retail")
    print(f"Storage: {data['storage']['count']} lockers @ ${data['storage']['fee']}/mo")
    print(f"Commercial: {data['commercial']['sf']} SF @ ${data['commercial']['rate']}/SF")
    print(f"Vacancy: {data['vacancy_rate']:.1%} residential, {data['commercial_vacancy']:.1%} commercial")
    print(f"Cap rates: {', '.join(f'{c:.2%}' for c in data['cap_rates'])}")
    if data.get('gfa'):
        print(f"GFA: {data['gfa']:,.0f} (from 1A)")
    else:
        print(f"GFA: will estimate from net rentable SF")

    # --- Interactive selections ---

    # 1. Municipality for DC rates
    municipalities = load_dc_rates()
    municipality = select_municipality(municipalities)

    # 2. Building type (mid-rise vs high-rise)
    building_type = select_building_type(data['total_units'])

    # --- Generate output ---
    project_name = data['address'].split(',')[0].strip().replace(' ', '_') or "project"
    project_name = re.sub(r'[^\w\-]', '', project_name)
    today = date.today().strftime("%Y%m%d")
    output_filename = f"Reverse_1B_{project_name}_{today}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    # Populate the template
    print(f"\nPopulating template...")
    log = populate_template(data, output_path, municipality=municipality, building_type=building_type)

    # Export project JSON for the presentation tool
    json_filename = f"Reverse_1B_{project_name}_{today}.json"
    json_path = os.path.join(OUTPUT_DIR, json_filename)
    export_project_json(data, json_path, municipality=municipality, building_type=building_type)
    print(f"Project JSON: {json_path}")

    # Save log
    log_filename = f"Reverse_1B_{project_name}_{today}_log.txt"
    log_path = os.path.join(OUTPUT_DIR, log_filename)
    with open(log_path, 'w') as f:
        f.write(f"SVN Rock — Reverse 1B Population Log\n")
        f.write(f"Source: {input_path}\n")
        f.write(f"Output: {output_path}\n")
        f.write(f"Municipality: {municipality['name'] if municipality else 'None selected'}\n")
        f.write(f"Building Type: {building_type}\n")
        f.write(f"Date: {date.today()}\n")
        f.write(f"\n")
        for line in log:
            f.write(line + "\n")

    print(f"\n{'='*60}")
    print(f"DONE!")
    print(f"{'='*60}")
    print(f"Output: {output_path}")
    print(f"Log:    {log_path}")
    if municipality:
        print(f"Municipality: {municipality['name']}")
        print(f"DC rates: 1-Bed ${municipality['rates']['1bed']:,} | "
              f"2-Bed ${municipality['rates']['2bed']:,} | "
              f"3-Bed ${municipality['rates']['3bed']:,}")
    print(f"Building type: {building_type}")
    print(f"\nLog contains {len([l for l in log if l.startswith('  ')])} cell writes "
          f"and {len([l for l in log if 'ESTIMATED' in l])} estimated values.")
    print(f"Review the log file for every assumption made.")

    # Show data freshness alerts if any datasets need attention
    alerts = get_alerts()
    if alerts:
        print(f"\n{'='*60}")
        print("DATA FRESHNESS ALERTS")
        print(f"{'='*60}")
        for a in alerts:
            icon = {"expired": "!!!", "warning": " ! ", "info": "   "}.get(a["level"], "   ")
            print(f"  [{icon}] {a['message']}")

    print(f"\nData sources: {get_data_sources_footer()}")


if __name__ == "__main__":
    main()
