"""
Compare Noor's manual Reverse 1B (the template) against our auto-generated output.
Both are for the Birchmount project, so values should match.

Usage:
    python compare_outputs.py                          # auto-finds latest Birchmount output
    python compare_outputs.py output/some_file.xlsx    # compare specific file
"""
import openpyxl
import sys
import glob
import os

# Find the auto-generated file
if len(sys.argv) > 1:
    auto_path = sys.argv[1]
else:
    # Find the most recent Birchmount output
    candidates = glob.glob("output/Reverse_1B_2240_Birchmount_Road_*.xlsx")
    candidates = [c for c in candidates if not c.startswith("~$")]
    if not candidates:
        print("No Birchmount output found in output/. Run populate_reverse_1b.py first.")
        sys.exit(1)
    auto_path = max(candidates, key=os.path.getmtime)

print(f"Template: reference/REVERSE_1B_Template.xlsx")
print(f"Auto:     {auto_path}")
print()

manual = openpyxl.load_workbook("reference/REVERSE_1B_Template.xlsx", data_only=True)
auto = openpyxl.load_workbook(auto_path, data_only=True)

# Sheets we care about (the ones we write to)
sheets_to_compare = ["1. 1A Proforma", "4. Area Schedule", "5. Key Assumptions"]

for sheet_name in sheets_to_compare:
    ws_m = manual[sheet_name]
    ws_a = auto[sheet_name]

    print(f"\n{'='*70}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*70}")

    matches = 0
    mismatches = 0
    only_manual = 0
    only_auto = 0

    # Get the max dimensions across both sheets
    max_row = max(ws_m.max_row, ws_a.max_row)
    max_col = max(ws_m.max_column, ws_a.max_column)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell_ref = ws_m.cell(row=row, column=col).coordinate
            val_m = ws_m.cell(row=row, column=col).value
            val_a = ws_a.cell(row=row, column=col).value

            # Skip if both empty
            if val_m is None and val_a is None:
                continue

            # Compare numerically when possible (handle float precision)
            if isinstance(val_m, (int, float)) and isinstance(val_a, (int, float)):
                if abs(val_m - val_a) < 0.01:
                    matches += 1
                    continue
                else:
                    mismatches += 1
                    print(f"  DIFF  {cell_ref:>6}: manual={val_m}  |  auto={val_a}")
                    continue

            # String comparison
            if val_m == val_a:
                matches += 1
            elif val_m is not None and val_a is None:
                only_manual += 1
                # Only show non-formula values missing from auto
                if not str(val_m).startswith("="):
                    print(f"  MANUAL ONLY  {cell_ref:>6}: {val_m}")
            elif val_m is None and val_a is not None:
                only_auto += 1
                print(f"  AUTO ONLY    {cell_ref:>6}: {val_a}")
            else:
                # Both have values but different
                # Skip formula differences (we loaded data_only=True so formulas show cached values)
                mismatches += 1
                print(f"  DIFF  {cell_ref:>6}: manual={val_m}  |  auto={val_a}")

    print(f"\n  Summary: {matches} match, {mismatches} different, {only_manual} manual-only, {only_auto} auto-only")

manual.close()
auto.close()
